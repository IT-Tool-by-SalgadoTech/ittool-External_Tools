#!/usr/bin/env python3
"""
IT Angel — Automated Defensive IT Inspection & Monitoring
By SalgadoTech — IT-Tool Project
v2.0.0
"""

import os, sys, time, socket, datetime, platform, threading, subprocess, shutil

# ─── Auto-install dependencies ────────────────────────────────────────────────
def install_if_missing(packages):
    import importlib
    for pkg, import_name in packages:
        try:
            importlib.import_module(import_name)
        except ImportError:
            print(f"  [+] Installing {pkg}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "--quiet"])

install_if_missing([
    ("psutil",   "psutil"),
    ("openpyxl", "openpyxl"),
    ("colorama", "colorama"),
])

import psutil, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from colorama import init, Fore, Style
init(autoreset=True)

# ─── Colors ───────────────────────────────────────────────────────────────────
RED    = Fore.RED    + Style.BRIGHT
GREEN  = Fore.GREEN  + Style.BRIGHT
YELLOW = Fore.YELLOW + Style.BRIGHT
CYAN   = Fore.CYAN   + Style.BRIGHT
BLUE   = Fore.BLUE   + Style.BRIGHT
WHITE  = Fore.WHITE  + Style.BRIGHT
MAGENTA= Fore.MAGENTA+ Style.BRIGHT
RESET  = Style.RESET_ALL
DIM    = Style.DIM

# ─── Excel styles ─────────────────────────────────────────────────────────────
FILL_RED    = PatternFill("solid", fgColor="C0392B")
FILL_YELLOW = PatternFill("solid", fgColor="F39C12")
FILL_GREEN  = PatternFill("solid", fgColor="27AE60")
FILL_HEADER = PatternFill("solid", fgColor="0A0F1E")
FILL_SUBHDR = PatternFill("solid", fgColor="1A2C5B")
FILL_ALT    = PatternFill("solid", fgColor="EAF0FB")
FILL_WHITE  = PatternFill("solid", fgColor="FFFFFF")
FILL_PURPLE = PatternFill("solid", fgColor="6C3483")
FILL_TEAL   = PatternFill("solid", fgColor="117A65")

FONT_TITLE      = Font(name="Calibri", bold=True,  color="FFFFFF", size=14)
FONT_WHITE_BOLD = Font(name="Calibri", bold=True,  color="FFFFFF", size=11)
FONT_BLACK      = Font(name="Calibri",              color="000000", size=10)
FONT_WHITE      = Font(name="Calibri",              color="FFFFFF", size=10)
FONT_WHITE_B10  = Font(name="Calibri", bold=True,  color="FFFFFF", size=10)
FONT_BLACK_BOLD = Font(name="Calibri", bold=True,  color="000000", size=10)

thin_border = Border(
    left=Side(style='thin',   color="CCCCCC"),
    right=Side(style='thin',  color="CCCCCC"),
    top=Side(style='thin',    color="CCCCCC"),
    bottom=Side(style='thin', color="CCCCCC"),
)

# ─── Globals ──────────────────────────────────────────────────────────────────
VERSION        = "2.0.0"
g_os_target    = ""
g_duration     = 0        # seconds; 0 = indefinite
g_stop_event   = threading.Event()
g_excel_path   = ""
g_tshark_path  = ""       # resolved at startup

# Baselines
g_baseline_ports     = set()
g_baseline_processes = set()
g_baseline_services  = set()
g_baseline_usb       = set()
g_baseline_tasks     = set()

CYCLE_INTERVAL = 15 * 60  # 15 minutes

# ─── Thread-safe Excel lock ───────────────────────────────────────────────────
# openpyxl is NOT thread-safe for writes. All wb.save() calls MUST acquire
# this lock so tshark background threads and the main cycle never corrupt
# the file by writing simultaneously (main Linux bug fix).
g_excel_lock = threading.Lock()

# ─── BANNER ───────────────────────────────────────────────────────────────────
def print_banner():
    os.system('cls' if os.name == 'nt' else 'clear')
    print(CYAN + r"")
    print(CYAN + r"  ██╗████████╗      █████╗ ███╗   ██╗ ██████╗ ███████╗██╗     ")
    print(CYAN + r"  ██║╚══██╔══╝     ██╔══██╗████╗  ██║██╔════╝ ██╔════╝██║     ")
    print(CYAN + r"  ██║   ██║        ███████║██╔██╗ ██║██║  ███╗█████╗  ██║     ")
    print(CYAN + r"  ██║   ██║        ██╔══██║██║╚██╗██║██║   ██║██╔══╝  ██║     ")
    print(CYAN + r"  ██║   ██║        ██║  ██║██║ ╚████║╚██████╔╝███████╗███████╗")
    print(CYAN + r"  ╚═╝   ╚═╝        ╚═╝  ╚═╝╚═╝  ╚═══╝ ╚═════╝ ╚══════╝╚══════╝")
    print(CYAN + r"")
    print(WHITE  + f"  {'─'*68}")
    print(CYAN   + f"  {'IT Angel  v' + VERSION:^68}")
    print(WHITE  + f"  {'Automated Defensive IT Inspection & Monitoring':^68}")
    print(DIM    + f"  {'By SalgadoTech  —  IT-Tool Project':^68}")
    print(WHITE  + f"  {'─'*68}\n")

# ─── TSHARK DETECTION ─────────────────────────────────────────────────────────
def detect_tshark():
    global g_tshark_path
    candidates_win = [
        r"C:\Program Files\Wireshark\tshark.exe",
        r"C:\Program Files (x86)\Wireshark\tshark.exe",
    ]
    if os.name == 'nt':
        for c in candidates_win:
            if os.path.isfile(c):
                g_tshark_path = c
                return True
        found = shutil.which("tshark")
        if found:
            g_tshark_path = found
            return True
    else:
        found = shutil.which("tshark")
        if found:
            g_tshark_path = found
            return True
    return False


def get_tshark_interface():
    """
    Find the tshark interface index whose IP matches the machine's active IP.
    Uses 'tshark -D' to list interfaces then 'tshark -i N --list-interfaces'
    or falls back to matching the active local IP via psutil.
    """
    if not g_tshark_path:
        return "1"

    local_ip = get_local_ip()   # e.g. 192.168.50.4

    try:
        # Step 1: get tshark interface list
        r = subprocess.run([g_tshark_path, "-D"],
                           capture_output=True, text=True, timeout=8)
        raw_lines = (r.stdout + r.stderr).strip().splitlines()

        # Build list of (index, full_name)
        iface_list = []
        for line in raw_lines:
            line = line.strip()
            if not line:
                continue
            parts = line.split(". ", 1)
            if len(parts) == 2 and parts[0].strip().isdigit():
                iface_list.append((parts[0].strip(), parts[1].strip()))

        if not iface_list:
            return "1"

        # Step 2: for each interface try a 1-packet capture and check if we
        # get packets — but that's slow. Instead match by IP using psutil.
        # psutil gives us iface name → IP. tshark -D on Windows gives
        # "\Device\NPF_{GUID} (Friendly Name)". Extract the friendly name
        # in parentheses and compare against psutil names.

        # Build psutil map: friendly_name_lower → ip
        psutil_map = {}
        stats = psutil.net_if_stats()
        for iface_name, addrs in psutil.net_if_addrs().items():
            for a in addrs:
                if a.family == socket.AF_INET and a.address == local_ip:
                    psutil_map[iface_name.lower()] = a.address

        # Try to match tshark friendly name against psutil name
        for idx, full_name in iface_list:
            # Extract friendly name: "\Device\NPF_{...} (Wi-Fi)" → "wi-fi"
            friendly = full_name
            if "(" in full_name and full_name.endswith(")"):
                friendly = full_name[full_name.rfind("(")+1:-1]
            friendly_lower = friendly.lower()

            # Direct match against psutil iface that has our local IP
            for ps_name in psutil_map:
                if ps_name in friendly_lower or friendly_lower in ps_name:
                    return idx
                # Also try without spaces/dashes
                ps_clean = ps_name.replace("-","").replace(" ","")
                fr_clean = friendly_lower.replace("-","").replace(" ","")
                if ps_clean in fr_clean or fr_clean in ps_clean:
                    return idx

        # Step 3: fallback — prefer non-virtual adapters by name hints
        PREFER = ["wi-fi", "wifi", "wireless", "wlan", "ethernet",
                  "local area connection"]
        AVOID  = ["loopback", "npcap loopback", "bluetooth", "vmware",
                  "virtualbox", "hyper-v", "local area connection* "]
        for idx, full_name in iface_list:
            name_l = full_name.lower()
            if any(a in name_l for a in AVOID):
                continue
            if any(p in name_l for p in PREFER):
                return idx

        # Step 4: last resort — first non-loopback interface
        for idx, full_name in iface_list:
            if "loopback" not in full_name.lower():
                return idx

        return iface_list[0][0]

    except Exception:
        return "1"

# ─── OS SELECTION ─────────────────────────────────────────────────────────────
def select_os():
    global g_os_target
    print(WHITE + "  Which system will IT Angel protect today?\n")
    print(f"  {GREEN}[1]{RESET} Windows")
    print(f"  {GREEN}[2]{RESET} Linux\n")
    while True:
        c = input(f"  {CYAN}Select (1/2): {RESET}").strip()
        if c == "1":   g_os_target = "windows"; print(f"\n  {GREEN}✔ Windows mode.{RESET}\n"); break
        elif c == "2": g_os_target = "linux";   print(f"\n  {GREEN}✔ Linux mode.{RESET}\n");   break
        else:          print(f"  {RED}Enter 1 or 2.{RESET}")

# ─── DURATION SELECTION (new UX) ─────────────────────────────────────────────
def select_duration():
    global g_duration
    print(WHITE  + "  ┌─────────────────────────────────────────────────────┐")
    print(WHITE  + "  │   How many hours do you want IT Angel to protect     │")
    print(WHITE  + "  │   this system?  Enter a number of hours.             │")
    print(WHITE  + "  │                                                       │")
    print(YELLOW + "  │   ➤  Type a number  (e.g.  2  or  0.5)               │")
    print(YELLOW + "  │   ➤  Type  F        for a quick Fast Check            │")
    print(YELLOW + "  │   ➤  Type  I        for Indefinite protection         │")
    print(WHITE  + "  └─────────────────────────────────────────────────────┘\n")
    while True:
        raw = input(f"  {CYAN}Your choice: {RESET}").strip().upper()
        if raw == "F":
            g_duration = -1   # flag for fast mode
            print(f"\n  {MAGENTA}⚡ Fast Check mode selected.{RESET}\n")
            break
        elif raw == "I":
            g_duration = 0
            print(f"\n  {GREEN}✔ Indefinite protection. Press Ctrl+C anytime to stop.{RESET}\n")
            break
        else:
            try:
                h = float(raw)
                if h <= 0: raise ValueError
                g_duration = int(h * 3600)
                label = f"{h:.1f}".rstrip('0').rstrip('.') + " hour(s)"
                print(f"\n  {GREEN}✔ IT Angel will protect this system for {label}.{RESET}\n")
                break
            except ValueError:
                print(f"  {RED}Invalid input. Enter a number, F, or I.{RESET}")

# ─── HELPERS ──────────────────────────────────────────────────────────────────
def now_str():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def run_cmd(cmd, shell=True, timeout=15):
    try:
        r = subprocess.run(cmd, shell=shell, capture_output=True, text=True, timeout=timeout)
        return r.stdout.strip()
    except Exception:
        return ""

def get_desktop_path():
    if os.name == 'nt':
        return os.path.join(os.path.expanduser("~"), "Desktop")
    # Linux with sudo -i: HOME=/root, use real user Desktop via SUDO_USER
    sudo_user = os.environ.get("SUDO_USER")
    if sudo_user:
        for candidate in [f"/home/{sudo_user}/Desktop", f"/home/{sudo_user}"]:
            if os.path.exists(candidate):
                return candidate
    d = os.path.join(os.path.expanduser("~"), "Desktop")
    return d if os.path.exists(d) else os.path.expanduser("~")

# ─── EXCEL SETUP ──────────────────────────────────────────────────────────────
def setup_excel():
    global g_excel_path
    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    hostname = socket.gethostname()
    fname    = f"IT_Angel_{hostname}_{ts}.xlsx"
    g_excel_path = os.path.join(get_desktop_path(), fname)

    wb      = openpyxl.Workbook()
    ws_tech = wb.active
    ws_tech.title = "Technical Log"
    ws_tech.sheet_view.showGridLines = False

    # ── Technical Log header ───────────────────────────────────────────────────
    ws_tech.merge_cells("A1:J1")
    c = ws_tech["A1"]
    c.value     = f"IT ANGEL — TECHNICAL LOG  |  Host: {hostname}  |  OS: {g_os_target.upper()}  |  Started: {now_str()}"
    c.fill      = FILL_HEADER
    c.font      = FONT_TITLE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_tech.row_dimensions[1].height = 30

    headers = ["Timestamp","Cycle","Category","Check","Finding","Details","Status","Delta","Severity","Recommendation"]
    for col, h in enumerate(headers, 1):
        cell = ws_tech.cell(row=2, column=col, value=h)
        cell.fill      = FILL_SUBHDR
        cell.font      = FONT_WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border
    ws_tech.row_dimensions[2].height = 22
    for i, w in enumerate([20,8,18,28,35,45,12,10,12,35], 1):
        ws_tech.column_dimensions[get_column_letter(i)].width = w
    ws_tech.freeze_panes = "A3"

    # ── Executive Summary ──────────────────────────────────────────────────────
    ws_exec = wb.create_sheet("Executive Summary")
    ws_exec.sheet_view.showGridLines = False
    ws_exec.merge_cells("A1:F1")
    e = ws_exec["A1"]
    e.value     = f"IT ANGEL — EXECUTIVE SUMMARY  |  {hostname}  |  {g_os_target.upper()}"
    e.fill      = FILL_HEADER
    e.font      = FONT_TITLE
    e.alignment = Alignment(horizontal="center", vertical="center")
    ws_exec.row_dimensions[1].height = 30
    for col, h in enumerate(["Timestamp","Cycle","Category","Status","Critical Findings","Recommendation"], 1):
        cell = ws_exec.cell(row=2, column=col, value=h)
        cell.fill      = FILL_SUBHDR
        cell.font      = FONT_WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
    ws_exec.row_dimensions[2].height = 22
    for i, w in enumerate([20,8,22,14,50,40], 1):
        ws_exec.column_dimensions[get_column_letter(i)].width = w
    ws_exec.freeze_panes = "A3"

    # ── Network Traffic sheet ──────────────────────────────────────────────────
    ws_net = wb.create_sheet("Network Traffic")
    ws_net.sheet_view.showGridLines = False
    ws_net.merge_cells("A1:F1")
    n = ws_net["A1"]
    n.value     = f"IT ANGEL — NETWORK TRAFFIC LOG  |  {hostname}  |  {g_os_target.upper()}"
    n.fill      = FILL_PURPLE
    n.font      = FONT_TITLE
    n.alignment = Alignment(horizontal="center", vertical="center")
    ws_net.row_dimensions[1].height = 30
    for col, h in enumerate(["Timestamp","Cycle","Source","Destination","Protocol","Info"], 1):
        cell = ws_net.cell(row=2, column=col, value=h)
        cell.fill      = FILL_TEAL
        cell.font      = FONT_WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
    ws_net.row_dimensions[2].height = 22
    for i, w in enumerate([20,8,22,22,14,50], 1):
        ws_net.column_dimensions[get_column_letter(i)].width = w
    ws_net.freeze_panes = "A3"

    wb.save(g_excel_path)
    print(f"  {GREEN}✔ Report file created:{RESET} {g_excel_path}\n")

# ─── EXCEL APPEND ─────────────────────────────────────────────────────────────
def append_to_excel(tech_rows, exec_rows, traffic_rows=None):
    with g_excel_lock:
        try:
            wb      = openpyxl.load_workbook(g_excel_path)
            ws_tech = wb["Technical Log"]
            ws_exec = wb["Executive Summary"]
            ws_net  = wb["Network Traffic"]

            # Technical Log
            next_r = ws_tech.max_row + 1
            for i, row in enumerate(tech_rows):
                r   = next_r + i
                alt = (r % 2 == 0)
                sev = row.get("Severity","").upper()
                vals = [row.get("Timestamp",""), row.get("Cycle",""), row.get("Category",""),
                        row.get("Check",""),     row.get("Finding",""), row.get("Details",""),
                        row.get("Status",""),    row.get("Delta",""),   row.get("Severity",""),
                        row.get("Recommendation","")]
                for col, val in enumerate(vals, 1):
                    cell = ws_tech.cell(row=r, column=col, value=val)
                    cell.border    = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if sev == "CRITICAL":
                        cell.fill = FILL_RED;    cell.font = FONT_WHITE_B10
                    elif sev == "WARNING":
                        cell.fill = FILL_YELLOW; cell.font = FONT_BLACK_BOLD
                    elif sev == "INFO":
                        cell.fill = FILL_ALT;    cell.font = FONT_BLACK
                    else:
                        cell.fill = FILL_ALT if alt else FILL_WHITE
                        cell.font = FONT_BLACK
                ws_tech.row_dimensions[r].height = 18

            # Executive Summary
            next_e = ws_exec.max_row + 1
            for i, row in enumerate(exec_rows):
                r   = next_e + i
                sev = row.get("Status","").upper()
                vals = [row.get("Timestamp",""), row.get("Cycle",""), row.get("Category",""),
                        row.get("Status",""),    row.get("Critical Findings",""), row.get("Recommendation","")]
                for col, val in enumerate(vals, 1):
                    cell = ws_exec.cell(row=r, column=col, value=val)
                    cell.border    = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if sev == "CRITICAL":
                        cell.fill = FILL_RED;    cell.font = FONT_WHITE_B10
                    elif sev == "WARNING":
                        cell.fill = FILL_YELLOW; cell.font = FONT_BLACK_BOLD
                    elif sev == "OK":
                        cell.fill = FILL_GREEN;  cell.font = FONT_WHITE
                    else:
                        cell.fill = FILL_WHITE;  cell.font = FONT_BLACK
                ws_exec.row_dimensions[r].height = 18

            # Network Traffic
            if traffic_rows:
                next_n = ws_net.max_row + 1
                for i, row in enumerate(traffic_rows):
                    r = next_n + i
                    alt = (r % 2 == 0)
                    vals = [row.get("Timestamp",""), row.get("Cycle",""), row.get("Source",""),
                            row.get("Destination",""), row.get("Protocol",""), row.get("Info","")]
                    for col, val in enumerate(vals, 1):
                        cell = ws_net.cell(row=r, column=col, value=val)
                        cell.border    = thin_border
                        cell.alignment = Alignment(vertical="center", wrap_text=True)
                        cell.fill      = FILL_ALT if alt else FILL_WHITE
                        cell.font      = FONT_BLACK
                    ws_net.row_dimensions[r].height = 16

            wb.save(g_excel_path)
        except Exception as ex:
            print(f"  {RED}[Excel Error] {ex}{RESET}")

# ═══════════════════════════════════════════════════════════════════════════════
#  CHECK MODULES
# ═══════════════════════════════════════════════════════════════════════════════

def check_system_snapshot():
    ts = now_str(); rows_t = []; rows_e = []
    hostname = socket.gethostname()
    ip       = get_local_ip()   # uses the proper APIPA-filtered function
    uname      = platform.uname()
    uptime_sec = time.time() - psutil.boot_time()
    uptime_str = str(datetime.timedelta(seconds=int(uptime_sec)))
    user       = os.environ.get("USERNAME") or os.environ.get("USER") or "N/A"
    ram_gb     = round(psutil.virtual_memory().total / (1024**3), 2)

    for check, finding in [
        ("Hostname",    hostname),
        ("IP Address",  ip),
        ("OS",          f"{uname.system} {uname.release} {uname.machine}"),
        ("Current User",user),
        ("Uptime",      uptime_str),
        ("CPU Cores",   str(psutil.cpu_count(logical=True))),
        ("RAM",         f"{ram_gb} GB"),
    ]:
        rows_t.append({"Timestamp":ts,"Cycle":"INIT","Category":"System Snapshot",
                        "Check":check,"Finding":finding,"Details":"",
                        "Status":"OK","Delta":"BASELINE","Severity":"INFO","Recommendation":""})
    rows_e.append({"Timestamp":ts,"Cycle":"INIT","Category":"System Snapshot","Status":"OK",
                   "Critical Findings":f"Host:{hostname} | IP:{ip} | User:{user} | Uptime:{uptime_str}",
                   "Recommendation":"Baseline captured."})
    return rows_t, rows_e


def check_network_adapters():
    ts = now_str(); rows_t = []; rows_e = []
    addrs = psutil.net_if_addrs(); stats = psutil.net_if_stats()
    for iface, addr_list in addrs.items():
        for addr in addr_list:
            if addr.family == socket.AF_INET:
                up  = stats[iface].isup if iface in stats else False
                sev = "OK" if up else "WARNING"
                rows_t.append({"Timestamp":ts,"Cycle":"INIT","Category":"Network Adapters",
                                "Check":iface,"Finding":addr.address,
                                "Details":f"Netmask:{addr.netmask} | Up:{up}",
                                "Status":"UP" if up else "DOWN","Delta":"BASELINE",
                                "Severity":sev,
                                "Recommendation":"" if up else f"Interface {iface} is DOWN"})
    findings = ", ".join([f"{i}:{a.address}" for i,al in addrs.items() for a in al if a.family==socket.AF_INET])
    rows_e.append({"Timestamp":ts,"Cycle":"INIT","Category":"Network Adapters","Status":"OK",
                   "Critical Findings":findings,"Recommendation":"Adapters catalogued."})
    return rows_t, rows_e


def check_active_connections(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    SUSPICIOUS = {4444,1337,31337,6666,9999,5555,8888,1234,65535,31338,12345,54321}
    total = 0
    try:
        seen = set()
        for c in psutil.net_connections(kind='inet'):
            if c.status not in ('ESTABLISHED','LISTEN'): continue
            try:    pname = psutil.Process(c.pid).name()
            except: pname = "N/A"
            laddr = f"{c.laddr.ip}:{c.laddr.port}" if c.laddr else ""
            raddr = f"{c.raddr.ip}:{c.raddr.port}" if c.raddr else "—"
            key   = (laddr, raddr, pname)
            if key in seen: continue
            seen.add(key); total += 1
            port = c.laddr.port if c.laddr else 0
            sev  = "OK"; rec = ""
            if port in SUSPICIOUS or (c.raddr and c.raddr.port in SUSPICIOUS):
                sev = "CRITICAL"
                rec = f"Suspicious port {port} — verify {pname}"
                critical_findings.append(f"SUSPICIOUS PORT {port} by {pname}")
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Active Connections",
                                "Check":f"{c.status} {laddr}","Finding":f"→ {raddr}",
                                "Details":f"Process:{pname} PID:{c.pid}",
                                "Status":c.status,"Delta":"","Severity":sev,"Recommendation":rec})
    except Exception as ex:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Active Connections",
                        "Check":"Error","Finding":str(ex),"Details":"",
                        "Status":"ERROR","Delta":"","Severity":"WARNING",
                        "Recommendation":"Run as administrator/root"})
    # Always add a summary row
    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Active Connections",
                    "Check":"Summary","Finding":f"{total} connections active",
                    "Details":"Only suspicious connections logged above",
                    "Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
    status = "CRITICAL" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Active Connections","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else f"{total} connections — none suspicious",
                   "Recommendation":"Investigate flagged ports" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_listening_ports(cycle):
    global g_baseline_ports
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    HIGH_RISK = {4444,1337,31337,6666,9999,5555,8888,23,512,513,514}
    current = set()
    try:
        for c in psutil.net_connections(kind='inet'):
            if c.status == 'LISTEN' and c.laddr:
                current.add(c.laddr.port)
    except: pass
    is_baseline = len(g_baseline_ports) == 0
    new_ports   = current - g_baseline_ports
    gone_ports  = g_baseline_ports - current

    # Always check high-risk ports
    for p in current:
        if p in HIGH_RISK:
            critical_findings.append(f"HIGH-RISK PORT: {p}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Listening Ports",
                            "Check":f"Port {p}","Finding":"HIGH-RISK OPEN","Details":"",
                            "Status":"CRITICAL","Delta":"","Severity":"CRITICAL",
                            "Recommendation":f"High-risk port {p} — verify immediately"})

    if is_baseline:
        # Baseline: just a summary
        g_baseline_ports = current
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Listening Ports",
                        "Check":"Baseline","Finding":f"{len(current)} ports catalogued",
                        "Details":f"Ports: {sorted(current)}",
                        "Status":"OK","Delta":"BASELINE","Severity":"INFO","Recommendation":""})
    else:
        # Only report new and gone
        for p in new_ports:
            sev = "CRITICAL" if p in HIGH_RISK else "WARNING"
            if p not in HIGH_RISK:  # HIGH_RISK already added above
                critical_findings.append(f"NEW PORT: {p}")
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Listening Ports",
                                "Check":f"Port {p}","Finding":"NEW — appeared this cycle",
                                "Details":"","Status":"LISTEN","Delta":"NEW ▲",
                                "Severity":sev,"Recommendation":f"New listening port {p} — verify"})
        for p in gone_ports:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Listening Ports",
                            "Check":f"Port {p}","Finding":"CLOSED","Details":"Was open last cycle",
                            "Status":"CLOSED","Delta":"GONE ▼","Severity":"INFO",
                            "Recommendation":"Verify if expected"})
        if not new_ports and not gone_ports and not critical_findings:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Listening Ports",
                            "Check":"Status","Finding":f"{len(current)} ports — no changes",
                            "Details":"","Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
        g_baseline_ports = current

    status = "CRITICAL" if any(p in HIGH_RISK for p in current) else "WARNING" if new_ports else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Listening Ports","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else f"{len(current)} ports — no anomalies",
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_network_processes(cycle):
    global g_baseline_processes
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    # Exact names only — no substring matching to avoid false positives like OneDrive
    SUSPICIOUS_EXACT = {
        "nc.exe","nc","ncat","ncat.exe","netcat","netcat.exe",
        "meterpreter","mimikatz","mimikatz.exe","empire","cobalt",
        "metasploit","nmap","nmap.exe","masscan","masscan.exe",
        "psexec","psexec.exe","wce","wce.exe","pwdump","fgdump",
        "gsecdump","quarks-pwdump","procdump.exe"
    }
    is_baseline = len(g_baseline_processes) == 0
    current = set()
    try:
        pids = {c.pid for c in psutil.net_connections(kind='inet') if c.pid}
        for pid in pids:
            try:
                p     = psutil.Process(pid)
                pname = p.name()
                current.add(pname)
                # Check suspicious — exact match only
                is_suspicious = pname.lower() in SUSPICIOUS_EXACT
                is_new = pname not in g_baseline_processes and not is_baseline
                sev = "OK"; rec = ""
                if is_suspicious:
                    sev = "CRITICAL"
                    rec = f"{pname} is a known attack tool — investigate immediately"
                    critical_findings.append(f"ATTACK TOOL: {pname} PID:{pid}")
                    try:    cmd = " ".join(p.cmdline()[:4])
                    except: cmd = "N/A"
                    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Net Processes",
                                    "Check":pname,"Finding":f"PID {pid}","Details":f"CMD:{cmd}",
                                    "Status":"SUSPICIOUS","Delta":"","Severity":"CRITICAL","Recommendation":rec})
                elif is_new:
                    sev = "WARNING"
                    rec = f"{pname} newly appeared with network activity"
                    critical_findings.append(f"NEW NET PROCESS: {pname}")
                    try:    cmd = " ".join(p.cmdline()[:4])
                    except: cmd = "N/A"
                    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Net Processes",
                                    "Check":pname,"Finding":f"PID {pid}","Details":f"CMD:{cmd}",
                                    "Status":"NEW","Delta":"NEW ▲","Severity":"WARNING","Recommendation":rec})
            except: pass
    except: pass

    if is_baseline:
        g_baseline_processes = current
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Net Processes",
                        "Check":"Baseline","Finding":f"{len(current)} processes with network activity",
                        "Details":"New processes will be flagged in next cycles",
                        "Status":"OK","Delta":"BASELINE","Severity":"INFO","Recommendation":""})
    else:
        g_baseline_processes = current
        if not critical_findings:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Net Processes",
                            "Check":"Status","Finding":f"{len(current)} processes — no new or suspicious",
                            "Details":"","Status":"OK","Delta":"","Severity":"OK","Recommendation":""})

    status = "CRITICAL" if any("ATTACK" in f for f in critical_findings) else \
             "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Net Processes","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else f"{len(current)} processes — no threats",
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_resources(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    cpu  = psutil.cpu_percent(interval=1)
    ram  = psutil.virtual_memory()
    disk = psutil.disk_usage('/')
    for check, finding, is_crit, is_warn, rec in [
        ("CPU Usage",     f"{cpu}%",              cpu>90,         cpu>75,         "CPU >90% — check runaway processes"),
        ("RAM Usage",     f"{ram.percent}%",       ram.percent>90, ram.percent>75, "RAM >90% — investigate memory hogs"),
        ("Disk Usage",    f"{disk.percent}%",      disk.percent>95,disk.percent>85,"Disk almost full — clean up"),
        ("RAM Available", f"{round(ram.available/1024**3,2)} GB", False, False,   ""),
        ("Disk Free",     f"{round(disk.free/1024**3,2)} GB",     False, False,   ""),
    ]:
        sev = "CRITICAL" if is_crit else "WARNING" if is_warn else "OK"
        if is_crit: critical_findings.append(f"{check}:{finding}")
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Resources",
                        "Check":check,"Finding":finding,"Details":"",
                        "Status":sev,"Delta":"","Severity":sev,
                        "Recommendation":rec if (is_crit or is_warn) else ""})
    # Top 5 CPU processes
    try:
        top = sorted(psutil.process_iter(['name','cpu_percent']), key=lambda p: p.info['cpu_percent'] or 0, reverse=True)[:5]
        for p in top:
            if (p.info['cpu_percent'] or 0) > 0:
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Resources",
                                "Check":f"Top CPU: {p.info['name']}",
                                "Finding":f"{p.info['cpu_percent']}%","Details":"",
                                "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    except: pass
    status = "CRITICAL" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Resources","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else f"CPU:{cpu}% RAM:{ram.percent}% Disk:{disk.percent}%",
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_logged_users(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    users = psutil.users()
    for u in users:
        login_t    = datetime.datetime.fromtimestamp(u.started).strftime("%Y-%m-%d %H:%M:%S")
        suspicious = u.host not in ('','::1','localhost','127.0.0.1',':0') and bool(u.host)
        sev        = "WARNING" if suspicious else "OK"
        if suspicious: critical_findings.append(f"REMOTE LOGIN: {u.name} from {u.host}")
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Logged Users",
                        "Check":u.name,"Finding":u.terminal or "N/A",
                        "Details":f"Host:{u.host or 'local'} | Login:{login_t}",
                        "Status":"REMOTE" if suspicious else "LOCAL",
                        "Delta":"","Severity":sev,
                        "Recommendation":f"Verify remote session from {u.host}" if suspicious else ""})
    if not users:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Logged Users",
                        "Check":"No users","Finding":"N/A","Details":"",
                        "Status":"OK","Delta":"","Severity":"INFO","Recommendation":""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Logged Users","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else f"{len(users)} user(s) active",
                   "Recommendation":"Verify remote sessions" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_services(cycle):
    global g_baseline_services
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []; current = set()
    CRITICAL_WIN = {"WinDefend", "MpsSvc", "EventLog", "Spooler", "LanmanServer"}
    CRITICAL_LIN = {"ssh","ufw","firewalld","cron","rsyslog","auditd","fail2ban","networking"}
    is_baseline  = len(g_baseline_services) == 0

    if g_os_target == "windows":
        try:
            for svc in psutil.win_service_iter():
                try:
                    info       = svc.as_dict()
                    name       = info['name']
                    status_svc = info['status']
                    current.add(name)
                    # On baseline — just check critical ones are running
                    if is_baseline:
                        if name in CRITICAL_WIN and status_svc != "running":
                            critical_findings.append(f"SERVICE DOWN: {name}")
                            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                                            "Check":name,"Finding":status_svc,
                                            "Details":info.get('display_name',''),
                                            "Status":status_svc.upper(),"Delta":"BASELINE",
                                            "Severity":"CRITICAL",
                                            "Recommendation":f"Critical service {name} is {status_svc}"})
                    else:
                        is_new = name not in g_baseline_services
                        if name in CRITICAL_WIN and status_svc != "running":
                            critical_findings.append(f"SERVICE DOWN: {name}")
                            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                                            "Check":name,"Finding":status_svc,
                                            "Details":info.get('display_name',''),
                                            "Status":status_svc.upper(),"Delta":"",
                                            "Severity":"CRITICAL",
                                            "Recommendation":f"Critical service {name} is {status_svc}"})
                        elif is_new:
                            critical_findings.append(f"NEW SERVICE: {name}")
                            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                                            "Check":name,"Finding":status_svc,
                                            "Details":info.get('display_name',''),
                                            "Status":status_svc.upper(),"Delta":"NEW ▲",
                                            "Severity":"WARNING",
                                            "Recommendation":f"New service {name} — verify"})
                except: pass
            # Summary row on baseline
            if is_baseline:
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                                "Check":"Baseline","Finding":f"{len(current)} services catalogued",
                                "Details":"Only critical failures and NEW services reported going forward",
                                "Status":"OK","Delta":"BASELINE","Severity":"INFO","Recommendation":""})
        except Exception as ex:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                            "Check":"Error","Finding":str(ex),"Details":"",
                            "Status":"ERROR","Delta":"","Severity":"WARNING",
                            "Recommendation":"Run as administrator"})
    else:
        out = run_cmd("systemctl list-units --type=service --state=running --no-pager --plain 2>/dev/null | awk '{print $1}'")
        for line in out.splitlines():
            name = line.strip().replace(".service","")
            if not name: continue
            current.add(name)
            if not is_baseline:
                is_new = name not in g_baseline_services
                if is_new:
                    critical_findings.append(f"NEW SERVICE: {name}")
                    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                                    "Check":name,"Finding":"running","Details":"",
                                    "Status":"RUNNING","Delta":"NEW ▲",
                                    "Severity":"WARNING","Recommendation":f"New service {name} — verify"})
        if is_baseline:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                            "Check":"Baseline","Finding":f"{len(current)} services catalogued",
                            "Details":"","Status":"OK","Delta":"BASELINE",
                            "Severity":"INFO","Recommendation":""})
        # Check critical linux services — only report if the service EXISTS on this system
        # (is known to systemd) but is not running. Skip if it was never installed.
        for svc in CRITICAL_LIN:
            if svc not in current:
                exists = run_cmd(f"systemctl list-unit-files {svc}.service 2>/dev/null | grep -c {svc}", timeout=5)
                try: svc_exists = int(exists.strip()) > 0
                except: svc_exists = False
                if svc_exists:
                    out2 = run_cmd(f"systemctl is-active {svc} 2>/dev/null")
                    if "active" not in out2:
                        critical_findings.append(f"SERVICE DOWN: {svc}")
                        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                                        "Check":svc,"Finding":"NOT running","Details":"Service is installed but stopped",
                                        "Status":"STOPPED","Delta":"","Severity":"WARNING",
                                        "Recommendation":f"Service {svc} is installed but not running — verify"})

    if is_baseline:
        g_baseline_services = current
    else:
        gone = g_baseline_services - current
        for s in gone:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services",
                            "Check":s,"Finding":"STOPPED","Details":"Was running previously",
                            "Status":"STOPPED","Delta":"GONE ▼","Severity":"WARNING",
                            "Recommendation":f"Service {s} stopped — verify"})
        g_baseline_services = current

    status = "CRITICAL" if any("CRITICAL" in f for f in critical_findings) else \
             "WARNING" if critical_findings else "OK"
    findings_str = "; ".join(critical_findings) if critical_findings else \
                   (f"Baseline: {len(current)} services catalogued" if is_baseline else "No changes detected")
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Services","Status":status,
                   "Critical Findings":findings_str,
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_firewall(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        out = run_cmd("netsh advfirewall show allprofiles state")
        for line in out.splitlines():
            if "State" in line:
                state = line.split()[-1].strip()
                sev   = "OK" if state.upper() == "ON" else "CRITICAL"
                if sev == "CRITICAL": critical_findings.append(f"FIREWALL OFF: {line.strip()}")
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Firewall",
                                "Check":"Profile State","Finding":line.strip(),"Details":"",
                                "Status":state.upper(),"Delta":"","Severity":sev,
                                "Recommendation":"" if sev=="OK" else "Enable firewall immediately"})
    else:
        ufw_out      = run_cmd("ufw status 2>/dev/null")
        iptables_out = run_cmd("iptables -L -n 2>/dev/null | head -20")
        nft_out      = run_cmd("nft list ruleset 2>/dev/null | head -5")
        fw_active    = ("active" in ufw_out.lower() or
                        "ACCEPT" in iptables_out or
                        bool(nft_out.strip()))
        detail       = ufw_out[:200] if ufw_out else (nft_out[:200] if nft_out else iptables_out[:200])
        sev          = "OK" if fw_active else "WARNING"
        if not fw_active: critical_findings.append("NO ACTIVE FIREWALL DETECTED")
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Firewall",
                        "Check":"Firewall Status","Finding":"Active" if fw_active else "Not detected",
                        "Details":detail,
                        "Status":"OK" if fw_active else "WARNING","Delta":"","Severity":sev,
                        "Recommendation":"" if fw_active else "Enable ufw or configure iptables/nftables"})
    status = "CRITICAL" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Firewall","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "Firewall active",
                   "Recommendation":"Enable firewall" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_usb_devices(cycle):
    global g_baseline_usb
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []; current = set()
    if g_os_target == "windows":
        out = run_cmd("wmic logicaldisk get caption,description,drivetype 2>nul")
    else:
        out = run_cmd("lsusb 2>/dev/null")
    for line in out.splitlines()[1:]:
        if line.strip():
            line_clean = line.strip()
            if os.name != 'nt':
                low = line_clean.lower()
                if "root hub" in low or "host controller" in low:
                    continue
            current.add(line_clean)
    is_baseline = len(g_baseline_usb) == 0
    new_usb     = current - g_baseline_usb
    gone_usb    = g_baseline_usb - current
    for dev in current:
        is_new = dev in new_usb and not is_baseline
        sev    = "WARNING" if is_new else "OK"
        if is_new: critical_findings.append(f"NEW USB: {dev[:50]}")
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"USB Devices",
                        "Check":"Device","Finding":dev[:80],"Details":"",
                        "Status":"CONNECTED","Delta":"NEW ▲" if is_new else ("BASELINE" if is_baseline else ""),
                        "Severity":sev,"Recommendation":"Verify new USB device" if is_new else ""})
    for dev in gone_usb:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"USB Devices",
                        "Check":"Device","Finding":dev[:80],"Details":"Disconnected",
                        "Status":"DISCONNECTED","Delta":"GONE ▼","Severity":"INFO",
                        "Recommendation":"USB device removed"})
    if not current and not gone_usb:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"USB Devices",
                        "Check":"No devices","Finding":"None","Details":"",
                        "Status":"OK","Delta":"","Severity":"INFO","Recommendation":""})
    g_baseline_usb = current
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"USB Devices","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else f"{len(current)} device(s) — no changes",
                   "Recommendation":"Verify new USB" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_dns_gateway(cycle):
    ts = now_str(); rows_t = []; rows_e = []; findings = []
    if g_os_target == "windows":
        gw_out  = run_cmd("ipconfig | findstr /i \"Default Gateway\"")
        dns_out = run_cmd("ipconfig /all | findstr /i \"DNS Servers\"")
    else:
        gw_out  = run_cmd("ip route | grep default | awk '{print $3}'")
        dns_out = run_cmd("cat /etc/resolv.conf | grep nameserver")
    for check, out in [("Default Gateway", gw_out), ("DNS Servers", dns_out)]:
        val = out[:120] if out else "N/A"
        findings.append(f"{check}: {val[:40]}")
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"DNS & Gateway",
                        "Check":check,"Finding":val,"Details":"",
                        "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"DNS & Gateway","Status":"OK",
                   "Critical Findings":" | ".join(findings),
                   "Recommendation":"Verify DNS matches expected servers"})
    return rows_t, rows_e


def check_scheduled_tasks(cycle):
    global g_baseline_tasks
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []; current = set()
    is_baseline = len(g_baseline_tasks) == 0

    if g_os_target == "windows":
        out = run_cmd('schtasks /query /fo LIST 2>nul | findstr /i "TaskName"', timeout=20)
        for line in out.splitlines():
            if "TaskName:" in line:
                tname = line.split("TaskName:")[-1].strip()
                if not tname: continue
                current.add(tname)
                is_new = tname not in g_baseline_tasks and not is_baseline
                if is_new:
                    critical_findings.append(f"NEW TASK: {tname[:50]}")
                    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Scheduled Tasks",
                                    "Check":tname[:60],"Finding":"NEW — appeared this cycle",
                                    "Details":"","Status":"WARNING","Delta":"NEW ▲",
                                    "Severity":"WARNING","Recommendation":"Verify this new scheduled task"})
        # On baseline: just log a summary, not every single task
        if is_baseline:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Scheduled Tasks",
                            "Check":"Baseline","Finding":f"{len(current)} tasks catalogued",
                            "Details":"Baseline established — only NEW tasks reported going forward",
                            "Status":"OK","Delta":"BASELINE","Severity":"INFO","Recommendation":""})
    else:
        out = run_cmd("crontab -l 2>/dev/null; ls /etc/cron.d/ 2>/dev/null; ls /etc/cron.daily/ 2>/dev/null")
        for line in out.splitlines():
            line = line.strip()
            if not line or line.startswith("#"): continue
            current.add(line)
            is_new = line not in g_baseline_tasks and not is_baseline
            if is_new:
                critical_findings.append(f"NEW CRON: {line[:50]}")
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Scheduled Tasks",
                                "Check":"cron","Finding":line[:80],"Details":"",
                                "Status":"WARNING","Delta":"NEW ▲",
                                "Severity":"WARNING","Recommendation":"New cron entry — verify"})
        if is_baseline:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Scheduled Tasks",
                            "Check":"Baseline","Finding":f"{len(current)} cron entries catalogued",
                            "Details":"Baseline established","Status":"OK","Delta":"BASELINE",
                            "Severity":"INFO","Recommendation":""})

    if is_baseline:
        g_baseline_tasks = current
    else:
        # Detect removed tasks
        gone = g_baseline_tasks - current
        for t in gone:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Scheduled Tasks",
                            "Check":t[:60],"Finding":"REMOVED since last cycle",
                            "Details":"","Status":"WARNING","Delta":"GONE ▼",
                            "Severity":"WARNING","Recommendation":"Verify task removal is expected"})
        g_baseline_tasks = current

    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Scheduled Tasks","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else
                   (f"Baseline: {len(current)} tasks catalogued" if is_baseline else "No new tasks detected"),
                   "Recommendation":"Verify new tasks" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_local_network(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        out = run_cmd("arp -a")
    else:
        out = run_cmd("arp -n 2>/dev/null || ip neigh 2>/dev/null")

    current_devices = {}
    for line in out.splitlines():
        line = line.strip()
        if not line or "Interface" in line or "Address" in line: continue
        parts = line.split()
        if parts:
            ip = parts[0]
            if ip.count(".") == 3 and not ip.startswith("224.") and not ip.startswith("239."):
                current_devices[ip] = line[:80]

    # Detect new devices vs baseline
    if not hasattr(check_local_network, '_baseline'):
        check_local_network._baseline = set(current_devices.keys())
        new_ips = set()
    else:
        new_ips = set(current_devices.keys()) - check_local_network._baseline
        check_local_network._baseline = set(current_devices.keys())

    # Report new devices as WARNING
    for ip in new_ips:
        if not ip.startswith("169.254."):
            critical_findings.append(f"NEW DEVICE: {ip}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"LAN Devices",
                            "Check":"New Device","Finding":ip,
                            "Details":current_devices.get(ip,""),
                            "Status":"NEW","Delta":"NEW ▲","Severity":"WARNING",
                            "Recommendation":f"New device {ip} joined the network — verify"})

    # Always add a summary row
    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"LAN Devices",
                    "Check":"LAN Summary","Finding":f"{len(current_devices)} devices on network",
                    "Details":f"IPs: {', '.join(sorted(current_devices.keys())[:10])}",
                    "Status":"OK","Delta":"","Severity":"OK","Recommendation":""})

    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"LAN Devices","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else f"{len(current_devices)} devices — no new devices",
                   "Recommendation":"Verify new devices" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_system_events(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        for evtid, desc, sev in [
            ("4625","Failed Login Attempts","CRITICAL"),
            ("4720","New User Account Created","CRITICAL"),
            ("7045","New Service Installed","WARNING"),
            ("4648","Logon with Explicit Credentials","WARNING"),
            ("4688","New Process Created","INFO"),
        ]:
            out = run_cmd(
                f'wevtutil qe Security /q:"*[System[EventID={evtid}]]" /c:5 /rd:true /f:text 2>nul | findstr /i "Date"',
                timeout=12)
            count = len([l for l in out.splitlines() if l.strip()])
            if count > 0 and sev in ("CRITICAL","WARNING"):
                critical_findings.append(f"Event {evtid} ({desc}): {count} recent")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"System Events",
                            "Check":f"Event {evtid}","Finding":desc,
                            "Details":f"Recent: {count}",
                            "Status":"FOUND" if count > 0 else "CLEAN",
                            "Delta":"","Severity":sev if count > 0 else "OK",
                            "Recommendation":f"Review event {evtid}" if count > 0 else ""})
    else:
        # Failed logins
        out = run_cmd("grep -i 'failed password\\|authentication failure' /var/log/auth.log 2>/dev/null | tail -10")
        count = len([l for l in out.splitlines() if l.strip()])
        if count > 0: critical_findings.append(f"Failed auth attempts: {count}")
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"System Events",
                        "Check":"Failed Logins","Finding":f"{count} recent",
                        "Details":out[:300] if out else "",
                        "Status":"FOUND" if count > 0 else "CLEAN","Delta":"",
                        "Severity":"CRITICAL" if count>5 else ("WARNING" if count>0 else "OK"),
                        "Recommendation":"Investigate brute force" if count>5 else ""})
        # Root sessions
        out2 = run_cmd("grep -i 'session opened for user root' /var/log/auth.log 2>/dev/null | tail -5")
        count2 = len([l for l in out2.splitlines() if l.strip()])
        if count2 > 0: critical_findings.append(f"Root sessions: {count2}")
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"System Events",
                        "Check":"Root Sessions","Finding":f"{count2} in log",
                        "Details":out2[:200] if out2 else "",
                        "Status":"FOUND" if count2>0 else "CLEAN","Delta":"",
                        "Severity":"WARNING" if count2>0 else "OK",
                        "Recommendation":"Verify root access is authorized" if count2>0 else ""})
        # Sudo usage
        out3 = run_cmd("grep -i 'sudo:' /var/log/auth.log 2>/dev/null | tail -5")
        count3 = len([l for l in out3.splitlines() if l.strip()])
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"System Events",
                        "Check":"Sudo Usage","Finding":f"{count3} recent sudo events",
                        "Details":out3[:200] if out3 else "",
                        "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})

    status = "CRITICAL" if any("CRITICAL" in str(r.get("Severity")) for r in rows_t) else \
             "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"System Events","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No suspicious events",
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_open_shares(cycle):
    """Windows: check open network shares. Linux: check NFS/Samba exports."""
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        out = run_cmd("net share 2>nul")
        for line in out.splitlines()[2:]:
            if not line.strip() or "---" in line: continue
            # Skip status lines like "The command completed successfully."
            if line.strip().lower().startswith("the ") or line.strip().lower().startswith("the	"):
                continue
            parts = line.split()
            if not parts: continue
            name = parts[0]
            # Valid share names don't contain spaces and are not full sentences
            if len(name) > 40 or " " in name or name.lower() in ("the","this","there"):
                continue
            sev  = "WARNING" if name.upper() not in ("C$","IPC$","ADMIN$","PRINT$") else "INFO"
            if sev == "WARNING": critical_findings.append(f"OPEN SHARE: {name}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Network Shares",
                            "Check":name,"Finding":line.strip()[:80],"Details":"",
                            "Status":"SHARED","Delta":"","Severity":sev,
                            "Recommendation":"Verify share is intentional" if sev=="WARNING" else ""})
    else:
        out = run_cmd("exportfs -v 2>/dev/null; smbstatus --shares 2>/dev/null | head -20")
        real_lines = []
        for line in out.splitlines():
            line = line.strip()
            if not line: continue
            if (line.startswith("-") or line.startswith("=") or
                    line.startswith("Service") or line.startswith("Samba") or
                    line.startswith("PID") or line.startswith("No ")):
                continue
            real_lines.append(line)
        if real_lines:
            for line in real_lines:
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Network Shares",
                                "Check":"Export/Share","Finding":line[:80],"Details":"",
                                "Status":"SHARED","Delta":"","Severity":"INFO","Recommendation":""})
        else:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Network Shares",
                            "Check":"Shares","Finding":"No active shares detected","Details":"",
                            "Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Network Shares","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No unexpected shares",
                   "Recommendation":"Verify shares" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_startup_items(cycle):
    """Windows: check startup registry & folder. Linux: check systemd enabled units."""
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        out = run_cmd('reg query "HKCU\\Software\\Microsoft\\Windows\\CurrentVersion\\Run" 2>nul')
        out += "\n"
        out += run_cmd('reg query "HKLM\\Software\\Microsoft\\Windows\\CurrentVersion\\Run" 2>nul')
        for line in out.splitlines():
            line = line.strip()
            if not line or line.startswith("HKEY"): continue
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Startup Items",
                            "Check":"Registry Run","Finding":line[:80],"Details":"",
                            "Status":"INFO","Delta":"","Severity":"INFO","Recommendation":""})
    else:
        out = run_cmd("systemctl list-unit-files --state=enabled --no-pager 2>/dev/null")
        current_units = set()
        for line in out.splitlines()[1:]:
            if not line.strip(): continue
            unit = line.split()[0] if line.split() else ""
            if unit: current_units.add(unit)
        if not hasattr(check_startup_items, '_baseline_lin'):
            check_startup_items._baseline_lin = set(current_units)
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Startup Items",
                            "Check":"Baseline","Finding":f"{len(current_units)} enabled units catalogued",
                            "Details":"","Status":"OK","Delta":"BASELINE","Severity":"INFO","Recommendation":""})
        else:
            new_units  = current_units - check_startup_items._baseline_lin
            gone_units = check_startup_items._baseline_lin - current_units
            for u in new_units:
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Startup Items",
                                "Check":"New Unit","Finding":u[:80],"Details":"",
                                "Status":"WARNING","Delta":"NEW ▲","Severity":"WARNING",
                                "Recommendation":f"New enabled unit: {u}"})
            for u in gone_units:
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Startup Items",
                                "Check":"Removed Unit","Finding":u[:80],"Details":"",
                                "Status":"INFO","Delta":"GONE ▼","Severity":"INFO","Recommendation":""})
            if not new_units and not gone_units:
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Startup Items",
                                "Check":"Status","Finding":"No changes","Details":"",
                                "Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
            check_startup_items._baseline_lin = current_units
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Startup Items","Status":"OK",
                   "Critical Findings":"Startup items catalogued","Recommendation":"Review for unexpected entries"})
    return rows_t, rows_e


# ─── NEW PROFESSIONAL CHECK MODULES ───────────────────────────────────────────

def check_local_users(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        out = run_cmd("net user 2>nul")
        admins_out = run_cmd("net localgroup administrators 2>nul")
        admin_list = set()
        in_members = False
        for line in admins_out.splitlines():
            if "---" in line: in_members = True; continue
            if in_members and line.strip() and "The command" not in line:
                admin_list.add(line.strip().lower())
        users_found = []
        for line in out.splitlines():
            line = line.strip()
            if not line or "User accounts" in line or "---" in line or "The command" in line: continue
            for u in line.split():
                if u: users_found.append(u)
        for u in users_found:
            is_admin = u.lower() in admin_list
            sev = "WARNING" if is_admin and u.lower() not in ("administrator","admin") else "OK"
            if is_admin: critical_findings.append(f"ADMIN USER: {u}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Local Users",
                            "Check":u,"Finding":"Admin" if is_admin else "Standard",
                            "Details":"Member of Administrators group" if is_admin else "",
                            "Status":"ADMIN" if is_admin else "STANDARD","Delta":"","Severity":sev,
                            "Recommendation":f"Verify {u} requires admin rights" if is_admin else ""})
    else:
        out = run_cmd("getent passwd | awk -F: '$7 !~ /nologin|false/ {print $1\":\"$3\":\"$6\":\"$7}'")
        sudo_out = run_cmd("getent group sudo wheel 2>/dev/null | cut -d: -f4")
        sudo_users = {u.strip() for line in sudo_out.splitlines() for u in line.split(",") if u.strip()}
        for line in out.splitlines():
            parts = line.split(":")
            if len(parts) < 4: continue
            uname, uid, home, shell = parts[0], parts[1], parts[2], parts[3]
            try: uid_int = int(uid)
            except: uid_int = -1
            is_root = uid == "0"; is_sudo = uname in sudo_users
            # UID 1000+ are normal human users — being in sudo group is expected
            is_primary_user = uid_int >= 1000
            sev = "CRITICAL" if is_root and uname != "root" else ("WARNING" if is_sudo and not is_primary_user else "OK")
            if is_root and uname != "root": critical_findings.append(f"EXTRA ROOT: {uname}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Local Users",
                            "Check":uname,"Finding":f"UID:{uid}",
                            "Details":f"Shell:{shell} | Sudo:{is_sudo}",
                            "Status":"ROOT" if is_root else ("SUDO" if is_sudo else "NORMAL"),
                            "Delta":"","Severity":sev,
                            "Recommendation":f"Verify {uname} requires elevated access" if sev != "OK" else ""})
    status = "CRITICAL" if any("CRITICAL" in f for f in critical_findings) else "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Local Users","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No unexpected privileged users",
                   "Recommendation":"Audit admin accounts" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_rdp_sessions(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        out = run_cmd("query session 2>nul")
        for line in out.splitlines()[1:]:
            if not line.strip(): continue
            parts = line.split()
            if len(parts) >= 3:
                session_name = parts[0].strip(">")
                username     = parts[1] if len(parts) > 1 else "N/A"
                state        = parts[3] if len(parts) > 3 else "N/A"
                is_rdp = "rdp" in session_name.lower() or "tcp" in session_name.lower()
                sev = "WARNING" if is_rdp else "OK"
                if is_rdp: critical_findings.append(f"RDP SESSION: {username} on {session_name}")
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"RDP Sessions",
                                "Check":session_name,"Finding":username,
                                "Details":f"State:{state}","Status":state,"Delta":"","Severity":sev,
                                "Recommendation":"Verify RDP session is authorized" if is_rdp else ""})
        rdp_events = run_cmd(
            'wevtutil qe Security /q:"*[System[EventID=4624] and EventData[Data[@Name=\'LogonType\']'
            'and(Data=\'10\')]]" /c:5 /rd:true /f:text 2>nul | findstr /i "Account Name"', timeout=10)
        count = len([l for l in rdp_events.splitlines() if l.strip()])
        if count > 0:
            critical_findings.append(f"RDP logon events: {count}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"RDP Sessions",
                            "Check":"Event 4624 (RDP)","Finding":f"{count} remote logon events",
                            "Details":rdp_events[:200],"Status":"FOUND","Delta":"","Severity":"WARNING",
                            "Recommendation":"Review recent remote logon activity"})
    else:
        out = run_cmd("who 2>/dev/null | grep pts")
        for line in out.splitlines():
            if not line.strip(): continue
            parts = line.split()
            uname = parts[0] if parts else "N/A"
            from_ = parts[-1].strip("()") if parts else "N/A"
            is_remote = "." in from_ or ":" in from_
            sev = "WARNING" if is_remote else "OK"
            if is_remote: critical_findings.append(f"SSH SESSION: {uname} from {from_}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"RDP Sessions",
                            "Check":f"SSH {uname}","Finding":from_,
                            "Details":"","Status":"REMOTE" if is_remote else "LOCAL",
                            "Delta":"","Severity":sev,
                            "Recommendation":"Verify remote SSH session" if is_remote else ""})
        if not out.strip():
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"RDP Sessions",
                            "Check":"SSH Sessions","Finding":"None active","Details":"",
                            "Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"RDP Sessions","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No remote sessions",
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_privileged_processes(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    EXPECTED_SYSTEM = {
        # Core Windows kernel & session
        "System","System Idle Process","Memory Compression","Secure System","Registry",
        "smss.exe","csrss.exe","wininit.exe","winlogon.exe","services.exe","lsass.exe",
        "lsm.exe","svchost.exe","spoolsv.exe","dwm.exe","sihost.exe","ctfmon.exe",
        "taskhostw.exe","taskhost.exe","conhost.exe","dllhost.exe","fontdrvhost.exe",
        "RuntimeBroker.exe","ShellExperienceHost.exe","StartMenuExperienceHost.exe",
        "SearchHost.exe","SearchIndexer.exe","SearchProtocolHost.exe","SearchFilterHost.exe",
        "audiodg.exe","WUDFHost.exe","wlms.exe","msdtc.exe","vds.exe","vssvc.exe",
        "unsecapp.exe","wbengine.exe","werfault.exe","werfaultsecure.exe","wsappx.exe",
        "compattelrunner.exe","msiexec.exe","TiWorker.exe","wuauclt.exe","dismhost.exe",
        "TrustedInstaller.exe","MoUsoCoreWorker.exe","usoclient.exe","sedsvc.exe",
        "UserOOBEBroker.exe","TextInputHost.exe","ApplicationFrameHost.exe",
        "backgroundTaskHost.exe","smartscreen.exe","browser_broker.exe",
        "SystemSettingsBroker.exe","DataExchangeHost.exe","PresentationFontCache.exe",
        "NgcIso.exe","LsaIso.exe","SgrmBroker.exe","SgrmAgent.exe",
        "PhoneExperienceHost.exe","Widgets.exe","WidgetService.exe",
        "MusNotifyIcon.exe","uhssvc.exe","OfficeClickToRun.exe","AppVShNotify.exe",
        "WinStore.App.exe","HxTsr.exe","msedge.exe","MicrosoftEdgeUpdate.exe",
        # Windows Security & Defender
        "MsMpEng.exe","NisSrv.exe","MpCmdRun.exe","SecurityHealthService.exe",
        "SecurityHealthSystray.exe","MpCopyAccelerator.exe","MpDefenderCoreService.exe",
        "WmiPrvSE.exe","WmiApSrv.exe",
        # Network services (legitimate SYSTEM)
        "SMSvcHost.exe","sshd.exe","ssh-agent.exe","wlanext.exe","dasHost.exe",
        "wslservice.exe","wsl.exe","wslhost.exe","Locator.exe","sqlwriter.exe",
        "DSRHost.exe","NgcIso.exe","pservice.exe","Service.exe",
        # NVIDIA
        "NVDisplay.Container.exe","nvcontainer.exe","nvdisplay.container.exe",
        "NvOAWrapperCache.exe","nvtelemetry.exe",
        # Intel graphics & audio
        "igfxEM.exe","igfxHK.exe","igfxTray.exe","IntelCpHDCPSvc.exe",
        # Realtek audio
        "RtkAudioService64.exe","RtkAudUService64.exe","RAVBg64.exe","RTUWPSrvcMain.exe",
        # DAX (Dolby/audio)
        "DAX3API.exe",
        # Lenovo
        "LenovoUtilityService.exe","LenovoVantageService.exe",
        "LenovoVantage-(VantageCoreAddin).exe",
        "LenovoVantage-(LenovoGamingSystemAddin).exe",
        "LenovoVantage-(HardwareScanAddin).exe",
        # NordVPN
        "nordvpn-service.exe","NordUpdateService.exe","nordvpn.exe",
        # HP
        "HPPrintScanDoctorService.exe","HPNetworkComm.exe",
        # Brave browser
        "BraveCrashHandler.exe","BraveCrashHandler64.exe",
        # Gaming / streaming
        "parsecd.exe","FMService64.exe","GamingServices.exe","GamingServicesNet.exe",
        "XblGameSave.exe","XboxPcApp.exe","OVRServer_x64.exe","OVRServiceLauncher.exe",
        "OVRRedir.exe","OVRServiceLauncher.exe",
        # Hyper-V / virtualization
        "vmms.exe","vmwp.exe","hvhost.exe","vmcompute.exe",
    }
    EXPECTED_ROOT_LIN = {
        "systemd","kthreadd","ksoftirqd","kworker","rcu_sched","rcu_preempt",
        "migration","watchdog","cpuhp","idle_inject","kdevtmpfs","netns","kauditd",
        "khungtaskd","oom_reaper","writeback","kcompactd","ksmd","khugepaged",
        "cryptd","kintegrityd","kblockd","blkcg_punt_bio","tpm_dev_wq","ata_sff",
        "md","edac-poller","devfreq_wq","kswapd","ecryptfs-kthrea",
        "cfg80211","kstrp","zswap","charger_manager","usb-storage",
        "bioset","sshd","cron","crond","rsyslogd","agetty","NetworkManager",
        "wpa_supplicant","polkitd","accounts-daemon","systemd-journald",
        "systemd-udevd","systemd-logind","systemd-resolved","systemd-timesyncd",
        "systemd-networkd","systemd-journal","systemd-userdbd","systemd-userwor",
        "dbus-daemon","avahi-daemon","bluetoothd","udisksd",
        "packagekitd","thermald","irqbalance","rtkit-daemon","upowerd","colord",
        "gdm","gdm3","lightdm","Xorg","Xwayland","gvfsd","gvfs-udisks2",
        "snapd","dockerd","containerd","VBoxService","vmtoolsd","open-vm-tools",
        "mysqld","apache2","nginx","php-fpm","postgres","redis","mongod",
        "python3","python","bash","sh","dash","zsh","fish",
        "networkd-dispat","unattended-upgr","ModemManager","switcheroo-cont",
        "nvidia-smi","nvidia-persistenced","i2cdetect",
        "rustdesk","nordvpnd","openvpn","wireguard",
        # Hardware & IRQ kernel threads confirmed on this Kali system
        "pool_workqueue_release","rcu_tasks_kthread","rcu_tasks_rude_kthread",
        "rcu_tasks_trace_kthread","rcu_exp_gp_kthread_worker","kdamond.0",
        "hwrng","haveged","smartd","watchdogd","psimon","fusermount3",
        "card0-crtc0","card0-crtc1","card0-crtc2","card0-crtc3","card0-crtc4",
        "jbd2/sda5-8","jbd2/sda1-8","jbd2/sda2-8","jbd2/nvme0n1p2-8",
        "scsi_eh_0","scsi_eh_1","scsi_eh_2","scsi_eh_3",
        # Transient root processes (normal system operation)
        "sudo","grep","ps","awk","sh","cat","sed","cut","sort","tr",
        "head","tail","find","ip","ss","systemctl","journalctl",
        "iptables","nft","ufw","loginctl","hostnamectl","apt","dpkg",
        # IT Angel own processes
        "tshark","dumpcap",
        # Bluetooth
        "krfcommd","obexd",
        # XFCE desktop session and components
        "xfce4-session","xfce4-panel","xfdesktop","xfwm4","xfsettingsd",
        "xfce4-notifyd","xfce4-power-man","xfce4-screensa","xfce4-terminal",
        "xfce4-appfinde","xfconfd","tumblerd","thunar",
        # XDG portals (ps truncates names to 15 chars)
        "xdg-desktop-por","xdg-desktop-portal",
        "xdg-permission-","xdg-permission-store",
        "xdg-document-po","xdg-document-portal",
        # GVFS virtual filesystem daemons
        "gvfsd-fuse","gvfsd-trash","gvfsd-network","gvfsd-computer",
        "gvfsd-metadata","gvfsd-recent",
        "gvfs-udisks2-vo","gvfs-udisks2-volume-monitor",
        "gvfs-gphoto2-vo","gvfs-gphoto2-volume-monitor",
        "gvfs-afc-volume","gvfs-afc-volume-monitor",
        "gvfs-goa-volume","gvfs-goa-volume-monitor",
        "gvfs-mtp-volume","gvfs-mtp-volume-monitor",
        # dconf settings daemon
        "dconf-service",
        # D-Bus
        "dbus-launch","dbus-run-session","dbus-broker","dbus-broker-lau",
        # Audio
        "pipewire","pipewire-pulse","wireplumber","pulseaudio",
        # Display/graphics
        "xrandr","xrdb","xsetroot","xset","xhost",
        # AT-SPI accessibility
        "at-spi-bus-laun","at-spi2-registr",
        # IBus input method
        "ibus-daemon","ibus-x11","ibus-portal","ibus-engine-sim",
        # GNOME keyring (used in XFCE too)
        "gnome-keyring-d",
        # PolicyKit
        "polkit",
        # UDisks / power / color
        "udisks2","upowerd","colord",
        # Geoclue location
        "geoclue",
        # fwupd firmware updater
        "fwupd",
        # CUPS printing
        "cupsd","cups-browsed",
        # Network
        "dhclient","dhcpcd",
        # Login managers
        "lightdm","slim","sddm",
        # ClamAV
        "clamd","freshclam",
        # RustDesk
        "scream",
    }
    KERNEL_THREAD_PREFIXES = (
        "kworker/","kthread","ksoftirqd/","migration/","watchdog/",
        "cpuhp/","idle_inject/","kcompactd","kswapd",
        "irq/","card0-crtc","jbd2/","kdamond.","scsi_eh_",
        "rcu_exp_","rcu_tasks_","pool_workqueue",
    )
    if g_os_target == "windows":
        out = run_cmd('tasklist /v /fo csv 2>nul', timeout=15)
        system_procs = []; unexpected = []
        for line in out.splitlines()[1:]:
            line = line.strip().strip('"')
            if not line: continue
            parts = [p.strip('"') for p in line.split('","')]
            if len(parts) < 7: continue
            pname = parts[0]; user = parts[6]
            if "SYSTEM" in user.upper() or "NT AUTHORITY" in user.upper():
                system_procs.append(pname)
                if pname not in EXPECTED_SYSTEM:
                    unexpected.append(pname)
                    critical_findings.append(f"UNEXPECTED SYSTEM: {pname}")
                    rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Privileged Processes",
                                    "Check":pname,"Finding":user[:40],
                                    "Details":"Not in known SYSTEM process list",
                                    "Status":"UNEXPECTED","Delta":"","Severity":"WARNING",
                                    "Recommendation":f"Investigate {pname} running as SYSTEM"})
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Privileged Processes",
                        "Check":"SYSTEM Summary","Finding":f"{len(system_procs)} total | {len(unexpected)} unexpected",
                        "Details":"Only unexpected processes listed above",
                        "Status":"WARNING" if unexpected else "OK","Delta":"",
                        "Severity":"WARNING" if unexpected else "OK",
                        "Recommendation":"Investigate" if unexpected else ""})
    else:
        out = run_cmd("ps -eo user,pid,comm --no-headers 2>/dev/null | grep '^root'")
        root_procs = []; unexpected = []
        for line in out.splitlines():
            parts = line.split()
            if len(parts) < 3: continue
            pname = parts[2]; pid = parts[1]
            root_procs.append(pname)
            is_kernel = any(pname.startswith(p) for p in KERNEL_THREAD_PREFIXES)
            if pname not in EXPECTED_ROOT_LIN and not is_kernel:
                unexpected.append(pname)
                critical_findings.append(f"UNEXPECTED ROOT: {pname}")
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Privileged Processes",
                                "Check":pname,"Finding":f"PID:{pid} User:root",
                                "Details":"Not in known root process list",
                                "Status":"UNEXPECTED","Delta":"","Severity":"WARNING",
                                "Recommendation":f"Investigate {pname} running as root"})
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Privileged Processes",
                        "Check":"Root Summary","Finding":f"{len(root_procs)} total | {len(unexpected)} unexpected",
                        "Details":"Only unexpected listed above",
                        "Status":"WARNING" if unexpected else "OK","Delta":"",
                        "Severity":"WARNING" if unexpected else "OK",
                        "Recommendation":"Investigate" if unexpected else ""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Privileged Processes","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No unexpected privileged processes",
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_recent_software(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        for reg_path in [
            r"HKLM\Software\Microsoft\Windows\CurrentVersion\Uninstall",
            r"HKLM\Software\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall",
        ]:
            out = run_cmd(f'reg query "{reg_path}" /s /v InstallDate 2>nul', timeout=20)
            for line in out.splitlines():
                if "InstallDate" not in line: continue
                parts = line.split(); date_str = parts[-1] if parts else ""
                if len(date_str) == 8 and date_str.isdigit():
                    try:
                        days_ago = (datetime.datetime.now() - datetime.datetime.strptime(date_str,"%Y%m%d")).days
                        if days_ago <= 7:
                            critical_findings.append(f"RECENT INSTALL ({days_ago}d ago): {date_str}")
                            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Recent Software",
                                            "Check":"Recent Install","Finding":date_str,
                                            "Details":f"Installed {days_ago} day(s) ago",
                                            "Status":"RECENT","Delta":"","Severity":"WARNING",
                                            "Recommendation":"Verify this installation is authorized"})
                    except: pass
        if not rows_t:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Recent Software",
                            "Check":"Recent Installs","Finding":"None in last 7 days",
                            "Details":"","Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
    else:
        out = run_cmd("grep 'install ' /var/log/dpkg.log 2>/dev/null | tail -10")
        if not out: out = run_cmd("grep 'Installed:' /var/log/yum.log 2>/dev/null | tail -10")
        for line in out.splitlines():
            if not line.strip(): continue
            critical_findings.append(f"PACKAGE: {line[:50]}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Recent Software",
                            "Check":"Package","Finding":line[:80],"Details":"",
                            "Status":"RECENT","Delta":"","Severity":"WARNING",
                            "Recommendation":"Verify this package is authorized"})
        if not rows_t:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Recent Software",
                            "Check":"Packages","Finding":"None detected",
                            "Details":"","Status":"OK","Delta":"","Severity":"OK","Recommendation":""})
    status = "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Recent Software","Status":status,
                   "Critical Findings":f"{len(critical_findings)} recent install(s)" if critical_findings else "No recent installs",
                   "Recommendation":"Audit recent software" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_system_file_integrity(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    import hashlib
    if g_os_target == "windows":
        critical_files = [
            r"C:\Windows\System32\drivers\etc\hosts",
            r"C:\Windows\System32\drivers\etc\services",
        ]
    else:
        critical_files = [
            "/etc/hosts","/etc/passwd","/etc/shadow",
            "/etc/sudoers","/etc/crontab","/etc/ssh/sshd_config",
        ]
    for fpath in critical_files:
        try:
            if not os.path.exists(fpath):
                rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"File Integrity",
                                "Check":os.path.basename(fpath),"Finding":"NOT FOUND","Details":"",
                                "Status":"MISSING","Delta":"","Severity":"WARNING",
                                "Recommendation":f"Expected file {fpath} is missing"})
                continue
            stat_r = os.stat(fpath)
            mtime  = datetime.datetime.fromtimestamp(stat_r.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            hours_ago = (time.time() - stat_r.st_mtime) / 3600
            with open(fpath,"rb") as f: chunk = f.read(4096)
            fhash = hashlib.md5(chunk).hexdigest()[:16]
            sev = "CRITICAL" if hours_ago < 1 else "WARNING" if hours_ago < 24 else "OK"
            if sev in ("CRITICAL","WARNING"):
                critical_findings.append(f"MODIFIED: {os.path.basename(fpath)} ({hours_ago:.1f}h ago)")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"File Integrity",
                            "Check":os.path.basename(fpath),"Finding":mtime,
                            "Details":f"Size:{stat_r.st_size}B | MD5:{fhash} | Modified:{hours_ago:.1f}h ago",
                            "Status":"MODIFIED" if sev!="OK" else "OK","Delta":"","Severity":sev,
                            "Recommendation":f"Verify change to {fpath}" if sev!="OK" else ""})
        except PermissionError:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"File Integrity",
                            "Check":os.path.basename(fpath),"Finding":"ACCESS DENIED","Details":"",
                            "Status":"DENIED","Delta":"","Severity":"INFO","Recommendation":""})
        except: pass
    status = "CRITICAL" if any("CRITICAL" in f for f in critical_findings) else \
             "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"File Integrity","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No critical file modifications",
                   "Recommendation":"Investigate modified system files" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_security_events_detail(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    if g_os_target == "windows":
        event_checks = [
            ("4625","Failed Login",            "CRITICAL", 3),
            ("4740","Account Lockout",          "CRITICAL", 1),
            ("4720","User Account Created",     "CRITICAL", 1),
            ("4726","User Account Deleted",     "CRITICAL", 1),
            ("4732","Added to Admin Group",     "CRITICAL", 1),
            ("4733","Removed from Admin Group", "WARNING",  1),
            ("4648","Explicit Credential Logon","WARNING",  1),
            ("4719","Audit Policy Changed",     "CRITICAL", 1),
            ("4698","Scheduled Task Created",   "WARNING",  1),
            ("4702","Scheduled Task Modified",  "WARNING",  1),
            ("7045","New Service Installed",    "WARNING",  1),
            ("1102","Audit Log Cleared",        "CRITICAL", 1),
        ]
        for evtid, desc, sev, threshold in event_checks:
            out = run_cmd(
                f'wevtutil qe Security /q:"*[System[EventID={evtid}]]" /c:{threshold+2} /rd:true /f:text 2>nul | findstr /i "Date"',
                timeout=10)
            count = len([l for l in out.splitlines() if l.strip()])
            if count >= threshold and sev != "INFO":
                critical_findings.append(f"Event {evtid} ({desc}): {count}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Security Events Detail",
                            "Check":f"Event {evtid}","Finding":desc,
                            "Details":f"Recent occurrences: {count}",
                            "Status":"FOUND" if count>0 else "CLEAN","Delta":"",
                            "Severity":sev if count>=threshold else "OK",
                            "Recommendation":f"Investigate Event {evtid}" if count>=threshold else ""})
    else:
        for cmd_str, desc, sev in [
            ("grep -c 'Failed password' /var/log/auth.log 2>/dev/null", "SSH Failed Passwords","CRITICAL"),
            ("grep -c 'Invalid user' /var/log/auth.log 2>/dev/null",    "Invalid SSH Users",   "WARNING"),
            ("grep -c 'Accepted password' /var/log/auth.log 2>/dev/null","SSH Successful Logins","INFO"),
            ("grep -c 'sudo:' /var/log/auth.log 2>/dev/null",           "Sudo Usage Events",   "INFO"),
            ("grep -c 'useradd\\|userdel\\|usermod' /var/log/auth.log 2>/dev/null","User Account Changes","CRITICAL"),
        ]:
            out = run_cmd(cmd_str, timeout=8).strip()
            try: count = int(out)
            except: count = 0
            if count > 0 and sev in ("CRITICAL","WARNING"):
                critical_findings.append(f"{desc}: {count}")
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Security Events Detail",
                            "Check":desc,"Finding":str(count),"Details":"",
                            "Status":"FOUND" if count>0 else "CLEAN","Delta":"",
                            "Severity":sev if count>0 else "OK",
                            "Recommendation":f"Review {desc}" if (count>0 and sev!="INFO") else ""})
    status = "CRITICAL" if any("CRITICAL" in f for f in critical_findings) else \
             "WARNING" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Security Events Detail","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No critical security events",
                   "Recommendation":"Investigate" if critical_findings else "Normal"})
    return rows_t, rows_e


def check_extended_ports(cycle):
    ts = now_str(); rows_t = []; rows_e = []; critical_findings = []
    BACKDOOR = {1337,4444,5555,6666,7777,8888,9999,31337,31338,12345,54321,65535,1234,2222,3333}
    RAT      = {1080,1099,3460,4899,5900,5901,5902,7000,8181,9001,9030,9050,9051}
    MINER    = {3333,4444,5555,7777,8888,9999,14444,14433,3004,3008,45560,45700}
    C2       = {6667,6668,6669,8080,8443,9090}
    try:
        flagged = {}
        for c in psutil.net_connections(kind='inet'):
            ports = set()
            if c.laddr: ports.add(c.laddr.port)
            if c.raddr: ports.add(c.raddr.port)
            for p in ports:
                cat = None
                if p in BACKDOOR: cat = "BACKDOOR"
                elif p in RAT:    cat = "RAT/REMOTE-ACCESS"
                elif p in MINER:  cat = "CRYPTO-MINER"
                elif p in C2:     cat = "C2/SUSPICIOUS"
                if cat:
                    try:    pname = psutil.Process(c.pid).name() if c.pid else "N/A"
                    except: pname = "N/A"
                    key = (p, pname)
                    if key not in flagged:
                        flagged[key] = cat
                        critical_findings.append(f"{cat} port {p} by {pname}")
                        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Extended Port Scan",
                                        "Check":f"Port {p}","Finding":cat,
                                        "Details":f"Process:{pname} | Status:{c.status}",
                                        "Status":c.status,"Delta":"","Severity":"CRITICAL",
                                        "Recommendation":f"INVESTIGATE: {cat} port {p} by {pname}"})
        if not flagged:
            rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Extended Port Scan",
                            "Check":"All ports","Finding":"No malicious ports detected","Details":"",
                            "Status":"CLEAN","Delta":"","Severity":"OK","Recommendation":""})
    except Exception as ex:
        rows_t.append({"Timestamp":ts,"Cycle":cycle,"Category":"Extended Port Scan",
                        "Check":"Error","Finding":str(ex),"Details":"",
                        "Status":"ERROR","Delta":"","Severity":"WARNING","Recommendation":"Run as administrator"})
    status = "CRITICAL" if critical_findings else "OK"
    rows_e.append({"Timestamp":ts,"Cycle":cycle,"Category":"Extended Port Scan","Status":status,
                   "Critical Findings":"; ".join(critical_findings) if critical_findings else "No malicious ports",
                   "Recommendation":"Investigate immediately" if critical_findings else "Normal"})
    return rows_t, rows_e

def get_local_ip():
    """Get primary local IP — filters out APIPA (169.254.x.x) and loopback."""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        if not ip.startswith("169.254.") and not ip.startswith("127."):
            return ip
    except Exception:
        pass
    # Fallback: scan psutil interfaces for a real routable IP
    try:
        for iface, addrs in psutil.net_if_addrs().items():
            stats = psutil.net_if_stats().get(iface)
            if not stats or not stats.isup:
                continue
            for a in addrs:
                if a.family == socket.AF_INET:
                    ip = a.address
                    if (not ip.startswith("127.") and
                            not ip.startswith("169.254.") and
                            not ip.startswith("0.")):
                        return ip
    except Exception:
        pass
    return "N/A"


def get_live_connections():
    """
    Read active network connections via psutil.
    Returns list sorted: ESTABLISHED first, then LISTEN, then rest.
    """
    results = []
    try:
        conns = psutil.net_connections(kind='inet')
        for c in conns:
            if c.status not in ('ESTABLISHED', 'LISTEN', 'CLOSE_WAIT', 'TIME_WAIT'):
                continue
            try:    pname = psutil.Process(c.pid).name() if c.pid else "System"
            except: pname = "System"
            src = f"{c.laddr.ip}:{c.laddr.port}" if c.laddr else ""
            dst = f"{c.raddr.ip}:{c.raddr.port}" if c.raddr else "—"
            results.append({
                "src":    src,
                "dst":    dst,
                "proto":  "TCP" if c.type == socket.SOCK_STREAM else "UDP",
                "status": c.status,
                "proc":   pname,
            })
    except Exception:
        pass
    # Sort: ESTABLISHED first
    order = {"ESTABLISHED": 0, "CLOSE_WAIT": 1, "TIME_WAIT": 2, "LISTEN": 3}
    results.sort(key=lambda x: (order.get(x["status"], 9), x["dst"]))
    return results[:20]


def get_lan_devices():
    """
    Quick ARP table scan — shows all devices currently visible on the LAN.
    Windows: arp -a + parallel ping sweep to refresh cache.
    Linux: ip neigh show (provides real MACs, no ping sweep needed).
    """
    devices = []
    seen    = set()
    local_ip = get_local_ip()

    if os.name == 'nt':
        # Windows: parallel ping sweep to refresh ARP cache
        subnet = ".".join(local_ip.split(".")[:3]) if local_ip and local_ip != "N/A" else ""
        alive_ips = set()
        def ping_one(ip):
            try:
                r = subprocess.run(["ping","-n","1","-w","300",ip], capture_output=True, timeout=1)
                if r.returncode == 0: alive_ips.add(ip)
            except: pass
        if subnet:
            threads = []
            for i in range(1, 255):
                ip = f"{subnet}.{i}"
                if ip == local_ip: alive_ips.add(ip); continue
                t = threading.Thread(target=ping_one, args=(ip,), daemon=True)
                threads.append(t); t.start()
            for t in threads: t.join(timeout=0.4)
        try:
            out = run_cmd("arp -a", timeout=3)
            for line in out.splitlines():
                line = line.strip()
                if not line or "Interface" in line or "Address" in line: continue
                parts = line.split()
                if len(parts) < 2: continue
                ip = parts[0]; mac = parts[1] if len(parts) > 1 else "?"
                if ip.count(".") != 3: continue
                if ip.startswith("224.") or ip.startswith("239."): continue
                if ip.endswith(".255") or ip.startswith("255."): continue
                if ip in seen: continue
                seen.add(ip)
                devices.append({"ip": ip, "mac": mac, "alive": ip in alive_ips or ip == local_ip})
        except: pass
    else:
        # Linux: ip neigh show — real MACs, no ping sweep
        try:
            out = run_cmd("ip neigh show 2>/dev/null", timeout=3)
            for line in out.splitlines():
                line = line.strip()
                if not line: continue
                parts = line.split()
                if not parts: continue
                ip = parts[0]
                if ip.count(".") != 3: continue
                if ip.startswith("224.") or ip.startswith("239."): continue
                if ip.endswith(".255") or ip.startswith("255."): continue
                if ip in seen: continue
                mac = "?"
                if "lladdr" in parts:
                    idx = parts.index("lladdr")
                    if idx + 1 < len(parts): mac = parts[idx + 1]
                state = parts[-1].upper() if parts else ""
                if state in ("FAILED", "INCOMPLETE"): continue
                seen.add(ip)
                devices.append({"ip": ip, "mac": mac, "alive": True})
        except: pass

    if local_ip and local_ip not in seen:
        devices.insert(0, {"ip": local_ip, "mac": "local", "alive": True})
    return devices


def get_net_stats():
    """Return bytes sent/recv for total interface."""
    try:
        s = psutil.net_io_counters()
        return s.bytes_sent, s.bytes_recv, s.packets_sent, s.packets_recv
    except Exception:
        return 0, 0, 0, 0


def get_local_subnet():
    """Return the first 3 octets of local IP to identify LAN range."""
    ip = get_local_ip()
    if ip and ip != "N/A":
        parts = ip.split(".")
        if len(parts) == 4:
            return ".".join(parts[:3])
    return "192.168.1"


def get_connection_bytes():
    """
    Get per-connection byte counters using psutil net_connections + proc io.
    Returns dict: dst_ip -> bytes_sent_to_it (approximate via process).
    This is used to detect ACTIVITY on existing connections.
    """
    activity = {}
    try:
        conns = psutil.net_connections(kind='inet')
        for c in conns:
            if c.status != 'ESTABLISHED' or not c.raddr:
                continue
            dst_ip = c.raddr.ip
            if dst_ip not in activity:
                activity[dst_ip] = 0
            # Count active connections per destination as a proxy for activity
            activity[dst_ip] += 1
    except Exception:
        pass
    return activity


def capture_tshark_background(cycle_num, duration_sec, result_holder, done_event):
    """
    Capture live traffic and parse it on-the-fly using tshark text output.

    Strategy (no .pcap file):
      - tshark writes tab-separated fields directly to stdout
      - We read line-by-line as they arrive (no post-processing delay)
      - Hard cap: at most MAX_PKTS packets or duration_sec seconds
      - Saves to Excel immediately when done, before the main thread
        starts the next cycle — eliminates the ~90s post-cycle lag.

    This replaces the old pcap→read two-step that caused:
      (a) ~60-90s tshark read/parse delay after the 15-min countdown
      (b) save_traffic_to_excel racing with the next cycle's append_to_excel
    """
    if not g_tshark_path:
        done_event.set()
        return

    MAX_PKTS = 500   # cap packets per cycle — keeps file lean & parse instant

    iface = get_tshark_interface()
    cmd = [
        g_tshark_path,
        "-i", iface,
        "-a", f"duration:{duration_sec}",
        "-c", str(MAX_PKTS),        # stop after MAX_PKTS packets
        "-T", "fields",
        "-e", "frame.time_epoch",
        "-e", "ip.src",
        "-e", "ip.dst",
        "-e", "_ws.col.Protocol",
        "-e", "_ws.col.Info",
        "-E", "separator=\t",
        "-E", "quote=n",
        "-E", "occurrence=f",
        "-l",                        # line-buffer stdout for live reading
    ]

    rows = []
    try:
        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,
            text=True,
            bufsize=1,           # line-buffered
        )
        try:
            for line in proc.stdout:
                if g_stop_event.is_set():
                    break
                line = line.rstrip()
                if not line:
                    continue
                parts = line.split("\t")
                try:
                    epoch = float(parts[0].strip())
                    ts_str = datetime.datetime.fromtimestamp(epoch).strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    ts_str = now_str()
                src   = parts[1].strip() if len(parts) > 1 else ""
                dst   = parts[2].strip() if len(parts) > 2 else ""
                proto = parts[3].strip() if len(parts) > 3 else ""
                info  = parts[4].strip()[:100] if len(parts) > 4 else ""
                if src or dst:
                    row = {
                        "Timestamp": ts_str, "Cycle": cycle_num,
                        "Source": src, "Destination": dst,
                        "Protocol": proto, "Info": info,
                    }
                    rows.append(row)
                    result_holder.append(row)
        except Exception:
            pass
        finally:
            try:
                proc.kill()
                proc.wait(timeout=5)
            except Exception:
                pass
    except Exception:
        pass

    # Save immediately — the lock guarantees no race with append_to_excel
    if rows:
        save_traffic_to_excel(rows)

    done_event.set()


def save_traffic_to_excel(traffic_rows):
    """Append traffic rows to the Network Traffic sheet — thread-safe."""
    if not traffic_rows:
        return
    with g_excel_lock:
        try:
            wb     = openpyxl.load_workbook(g_excel_path)
            ws_net = wb["Network Traffic"]
            next_n = ws_net.max_row + 1
            for i, row in enumerate(traffic_rows):
                r   = next_n + i
                alt = (r % 2 == 0)
                vals = [row.get("Timestamp",""), row.get("Cycle",""),
                        row.get("Source",""),     row.get("Destination",""),
                        row.get("Protocol",""),   row.get("Info","")]
                for col, val in enumerate(vals, 1):
                    cell           = ws_net.cell(row=r, column=col, value=val)
                    cell.border    = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.fill      = FILL_ALT if alt else FILL_WHITE
                    cell.font      = FONT_BLACK
                ws_net.row_dimensions[r].height = 16
            wb.save(g_excel_path)
        except Exception:
            pass


def show_live_traffic_and_countdown(wait_seconds, cycle_num, duration_end=None):
    """
    During the 15-min wait between cycles:
      1. Live connections via psutil — refreshes every 2s
      2. LAN Devices panel — ARP scan detects new devices
      3. tshark background capture — saves to Excel at end

    On Linux the display loop runs in a daemon thread so time.sleep()
    inside it cannot block the main thread. The main thread waits on a
    threading.Event — this eliminates the inter-cycle freeze on Kali.
    Windows: same behaviour, also runs in thread for consistency.
    """
    local_ip     = get_local_ip()
    local_subnet = get_local_subnet()
    iface        = get_tshark_interface()
    start_time   = time.time()

    # ── Start tshark background capture ───────────────────────────────────────
    tshark_rows = []
    tshark_done = threading.Event()
    if g_tshark_path:
        tshark_thread = threading.Thread(
            target=capture_tshark_background,
            args=(cycle_num, wait_seconds - 5, tshark_rows, tshark_done),
            daemon=True
        )
        tshark_thread.start()

    # ── LAN devices — cached and refreshed in background every 10s ────────────
    _lan_cache      = list(get_lan_devices())
    _lan_cache_lock = threading.Lock()

    def _refresh_lan_loop():
        nonlocal _lan_cache
        while not g_stop_event.is_set():
            for _ in range(5):          # wait 10s (5 x 2s)
                if g_stop_event.is_set():
                    return
                time.sleep(2)
            try:
                fresh = get_lan_devices()
                with _lan_cache_lock:
                    _lan_cache = fresh
            except Exception:
                pass

    lan_bg = threading.Thread(target=_refresh_lan_loop, daemon=True)
    lan_bg.start()

    def _get_lan_now():
        with _lan_cache_lock:
            return list(_lan_cache)

    # ── Event fired when countdown reaches 0 ──────────────────────────────────
    _display_done = threading.Event()

    # ── Display loop — runs in its own thread so main thread never sleeps ─────
    def _display_loop():
        prev_sent, prev_recv, _, _ = get_net_stats()
        baseline_lan  = {d["ip"] for d in _get_lan_now()}
        prev_conn_map = get_connection_bytes()
        activity_log  = {}
        REFRESH = 2

        try:
            while True:
                elapsed = time.time() - start_time
                left    = max(0, wait_seconds - int(elapsed))
                m, s    = divmod(left, 60)
                if g_stop_event.is_set() or left <= 0:
                    break

                conns        = get_live_connections()
                lan_devices  = _get_lan_now()
                cur_sent, cur_recv, _, _ = get_net_stats()
                delta_up     = max(0, (cur_sent - prev_sent)) // 1024
                delta_down   = max(0, (cur_recv - prev_recv)) // 1024
                prev_sent, prev_recv = cur_sent, cur_recv

                new_devices = [d for d in lan_devices
                               if d["ip"] not in baseline_lan
                               and not d["ip"].startswith("224.")
                               and not d["ip"].startswith("239.")
                               and not d["ip"].startswith("169.254.")]

                cur_conn_map = get_connection_bytes()
                active_lan   = {}
                for dst_ip, count in cur_conn_map.items():
                    if dst_ip.startswith(local_subnet + ".") and dst_ip != local_ip:
                        if count > 0:
                            activity_log[dst_ip] = activity_log.get(dst_ip, 0) + REFRESH
                            active_lan[dst_ip] = True
                        elif dst_ip in activity_log and activity_log[dst_ip] > 0:
                            activity_log[dst_ip] = max(0, activity_log[dst_ip] - REFRESH)
                prev_conn_map = cur_conn_map

                lines_printed = getattr(_display_loop, '_lines', 0)
                if lines_printed > 0:
                    print(f"\033[{lines_printed}A\033[J", end="", flush=True)

                frame_lines = 0
                def pr(text=""):
                    nonlocal frame_lines
                    print(text)
                    frame_lines += 1

                remain_label = ""
                if duration_end:
                    tl = max(0, int(duration_end - time.time()))
                    th, tm2 = divmod(tl, 3600); tm2, ts2 = divmod(tm2, 60)
                    remain_label = f"  {DIM}| Session ends: {th:02d}:{tm2:02d}:{ts2:02d}{RESET}"

                pr(f"  {CYAN}{chr(9472)*74}{RESET}")
                pr(f"  {WHITE}\u23f1  Next cycle in: {CYAN}{m:02d}:{s:02d}{RESET}  "
                   f"{GREEN}This machine: {local_ip}{RESET}{remain_label}")
                pr(f"  {WHITE}psutil live  "
                   + (f"| tshark \u2192 [{iface}]" if g_tshark_path else "| tshark: not installed")
                   + f"  {CYAN}\u2191{delta_up}KB  \u2193{delta_down}KB/s{RESET}")
                pr(f"  {CYAN}{chr(9472)*74}{RESET}")

                for dst_ip, secs in activity_log.items():
                    if secs > 0:
                        procs = [c["proc"] for c in conns if dst_ip in c.get("dst","")]
                        proc_label = procs[0] if procs else "System"
                        bar = "\u2588" * min(int(secs / 2), 20)
                        pr(f"  {MAGENTA}\u2605 ACTIVITY \u2192 {dst_ip:<18} "
                           f"via {proc_label:<14} [{bar}]{RESET}")

                pr(f"  {YELLOW}{'PROCESS':<15} {'SOURCE':<21} {'DESTINATION':<21} {'P':<4} {'STATUS'}{RESET}")
                pr(f"  {YELLOW}{chr(9472)*74}{RESET}")

                established = [c for c in conns if c["status"] == "ESTABLISHED"]
                listening   = [c for c in conns if c["status"] == "LISTEN"]
                other       = [c for c in conns if c["status"] not in ("ESTABLISHED","LISTEN")]
                shown = 0

                for c in (established + other + listening)[:14]:
                    proc   = c["proc"][:13]
                    src    = c["src"][:19]
                    dst    = c["dst"][:19]
                    proto  = c["proto"][:3]
                    status = c["status"][:10]
                    mine   = local_ip in c["src"]
                    dst_ip_only = c["dst"].split(":")[0] if c["dst"] else ""
                    is_lan_dst = dst_ip_only.startswith(local_subnet + ".")
                    is_active  = dst_ip_only in active_lan

                    if is_active and is_lan_dst:
                        color = MAGENTA; marker = f"{MAGENTA}\u2605{RESET} "
                    elif is_lan_dst and c["status"] == "ESTABLISHED":
                        color = YELLOW;  marker = f"{YELLOW}\u25ba{RESET} "
                    elif mine and c["status"] == "ESTABLISHED":
                        color = CYAN;    marker = f"{GREEN}\u25ba{RESET} "
                    elif c["status"] == "ESTABLISHED":
                        color = WHITE;   marker = "  "
                    elif c["status"] == "LISTEN":
                        color = WHITE;   marker = "  "
                    else:
                        color = YELLOW;  marker = "  "

                    pr(f"  {marker}{color}{proc:<15} {src:<21} {dst:<21} {proto:<4} {status}{RESET}")
                    shown += 1

                if shown == 0:
                    pr(f"  {WHITE}  No active connections{RESET}")

                pr(f"  {YELLOW}{chr(9472)*74}{RESET}")
                pr(f"  {WHITE}{len(established)} established  {len(listening)} listening  "
                   f"| {GREEN}\u25ba = this machine  {MAGENTA}\u2605 = LAN activity{RESET}  "
                   f"{WHITE}| refreshes every {REFRESH}s{RESET}")

                pr(f"  {CYAN}{chr(9472)*74}{RESET}")
                pr(f"  {WHITE}LAN Devices ({len(lan_devices)} on {local_subnet}.0/24):{RESET}")

                if new_devices:
                    for d in new_devices:
                        pr(f"  {RED}\u2605 NEW DEVICE DETECTED: {d['ip']:<18} MAC: {d['mac']}{RESET}")

                shown_ips = set()
                for d in lan_devices:
                    ip  = d["ip"]
                    mac = d["mac"]
                    if ip.startswith("169.254.") or ip.startswith("224.") or ip.startswith("239."):
                        continue
                    if ip in shown_ips:
                        continue
                    shown_ips.add(ip)
                    is_me       = (ip == local_ip)
                    is_active_d = ip in active_lan
                    is_new      = ip in {nd["ip"] for nd in new_devices}

                    if is_new:
                        color = RED;     marker = f"{RED}\u2605{RESET} "
                    elif is_active_d:
                        color = MAGENTA; marker = f"{MAGENTA}\u2605{RESET} "
                    elif is_me:
                        color = GREEN;   marker = f"{GREEN}\u25ba{RESET} "
                    else:
                        color = WHITE;   marker = "  "

                    pr(f"  {marker}{color}{ip:<18}{RESET}  {WHITE}{mac}{RESET}")

                pr(f"  {CYAN}{chr(9472)*74}{RESET}")
                elapsed_total = int(time.time() - start_time)
                eh, em2 = divmod(elapsed_total, 3600); em2, es = divmod(em2, 60)
                pr(f"  \u25cf Next cycle: {CYAN}{m:02d}:{s:02d}{RESET}"
                   + (f"  {DIM}Session ends: {remain_label.strip()}{RESET}" if remain_label else ""))

                _display_loop._lines = frame_lines
                time.sleep(REFRESH)

        except KeyboardInterrupt:
            g_stop_event.set()
        except Exception:
            pass
        finally:
            lines_printed = getattr(_display_loop, '_lines', 0)
            if lines_printed > 0:
                print(f"\033[{lines_printed}A\033[J", end="", flush=True)
            _display_loop._lines = 0
            _display_done.set()   # Always fire — guarantees main thread unblocks

    # ── Launch display in thread; main thread waits on Event (never sleeps) ───
    display_thread = threading.Thread(target=_display_loop, daemon=True)
    display_thread.start()

    # Poll with 1s timeout so Ctrl+C is caught quickly.
    # Hard timeout = wait_seconds + 30s as absolute safety net against freeze.
    deadline = time.time() + wait_seconds + 30
    while not _display_done.wait(timeout=1.0):
        if g_stop_event.is_set():
            break
        if time.time() >= deadline:
            # Safety: force the event so we never block the cycle loop
            _display_done.set()
            break

    # ── Wait for tshark to finish saving (should be nearly instant now) ─────────
    if g_tshark_path:
        # With streaming output, tshark finishes at the same time as the countdown.
        # Give it at most 5s to flush and save — it should be done already.
        tshark_done.wait(timeout=5)
        if tshark_rows:
            print(f"  {GREEN}\u2714 {len(tshark_rows)} packets \u2192 Network Traffic sheet.{RESET}")
        else:
            print(f"  {DIM}tshark: no packets captured this cycle.{RESET}")
    print()


# ─── FULL CYCLE RUNNER ────────────────────────────────────────────────────────
def run_full_cycle(cycle_num):
    """Run ALL checks every cycle."""
    all_tech = []; all_exec = []

    def run_check(label, fn, *args):
        print(f"    {DIM}[ {label} ]{RESET}", end="\r")
        try:
            t, e = fn(*args)
            all_tech.extend(t); all_exec.extend(e)
            crit = any(r.get("Severity") == "CRITICAL" for r in t)
            warn = any(r.get("Severity") == "WARNING"  for r in t)
            icon = f"{RED}●{RESET}" if crit else (f"{YELLOW}●{RESET}" if warn else f"{GREEN}●{RESET}")
            stat = "CRITICAL" if crit else ("WARNING" if warn else "OK")
            print(f"    {icon} {WHITE}{label:<35}{RESET} {stat}")
        except Exception as ex:
            print(f"    {RED}✗ {label}: {ex}{RESET}")

    print(f"\n  {CYAN}{'─'*60}{RESET}")
    print(f"  {WHITE}Cycle {cycle_num} — {now_str()}{RESET}")
    print(f"  {CYAN}{'─'*60}{RESET}\n")

    run_check("Active Connections",      check_active_connections,     cycle_num)
    run_check("Listening Ports",         check_listening_ports,        cycle_num)
    run_check("Extended Port Scan",      check_extended_ports,         cycle_num)
    run_check("Net Processes",           check_network_processes,      cycle_num)
    run_check("Privileged Processes",    check_privileged_processes,   cycle_num)
    run_check("Resources (CPU/RAM)",     check_resources,              cycle_num)
    run_check("Logged Users",            check_logged_users,           cycle_num)
    run_check("Local Users & Admins",    check_local_users,            cycle_num)
    run_check("RDP / Remote Sessions",   check_rdp_sessions,           cycle_num)
    run_check("Services",                check_services,               cycle_num)
    run_check("Firewall",                check_firewall,               cycle_num)
    run_check("USB Devices",             check_usb_devices,            cycle_num)
    run_check("DNS & Gateway",           check_dns_gateway,            cycle_num)
    run_check("Scheduled Tasks",         check_scheduled_tasks,        cycle_num)
    run_check("LAN Devices (ARP)",       check_local_network,          cycle_num)
    run_check("System Events",           check_system_events,          cycle_num)
    run_check("Security Events Detail",  check_security_events_detail, cycle_num)
    run_check("Network Shares",          check_open_shares,            cycle_num)
    run_check("Startup Items",           check_startup_items,          cycle_num)
    run_check("Recent Software",         check_recent_software,        cycle_num)
    run_check("File Integrity",          check_system_file_integrity,  cycle_num)

    print(f"\n  {DIM}Saving to Excel...{RESET}", end="\r")
    append_to_excel(all_tech, all_exec)
    print(f"  {GREEN}✔ Excel updated — {g_excel_path}{RESET}")
    return all_tech, all_exec


# ─── BASELINE ─────────────────────────────────────────────────────────────────
def run_baseline():
    """
    Captures initial state silently — no cycle data written.
    Sets all baseline globals so Cycle 1 detects real changes.
    """
    print(f"\n  {CYAN}Running initial system baseline...{RESET}\n")
    all_tech = []; all_exec = []

    # System info — these write to Excel as INIT rows
    for label, fn in [("System Snapshot",  check_system_snapshot),
                      ("Network Adapters", check_network_adapters)]:
        print(f"    {DIM}[ {label} ]{RESET}", end="\r")
        try:
            t, e = fn()
            all_tech.extend(t); all_exec.extend(e)
            print(f"    {GREEN}✔{RESET} {WHITE}{label}{RESET}")
        except Exception as ex:
            print(f"    {RED}✗ {label}: {ex}{RESET}")
    append_to_excel(all_tech, all_exec)

    # Silent baseline — populates all global sets, rows discarded
    silent = [
        ("Ports baseline",     check_listening_ports),
        ("Processes baseline", check_network_processes),
        ("Services baseline",  check_services),
        ("Tasks baseline",     check_scheduled_tasks),
        ("USB baseline",       check_usb_devices),
    ]
    for label, fn in silent:
        print(f"    {DIM}[ {label} ]{RESET}", end="\r")
        try:
            fn("BASELINE")
            print(f"    {GREEN}✔{RESET} {WHITE}{label}{RESET}")
        except Exception as ex:
            print(f"    {RED}✗ {label}: {ex}{RESET}")

    # Also warm up LAN baseline
    try: check_local_network("BASELINE")
    except: pass

    print(f"\n  {GREEN}✔ Baseline complete — Cycle 1 will be the first active check.{RESET}\n")


# ─── FAST CHECK MODE ──────────────────────────────────────────────────────────
def run_fast_check():
    print(f"\n  {MAGENTA}⚡ Fast Check — running all checks once...{RESET}\n")
    all_tech = []; all_exec = []
    for label, fn in [("System Snapshot", check_system_snapshot),
                      ("Network Adapters", check_network_adapters)]:
        try:
            t, e = fn()
            all_tech.extend(t); all_exec.extend(e)
            print(f"    {GREEN}✔{RESET} {WHITE}{label}{RESET}")
        except: pass
    append_to_excel(all_tech, all_exec)
    run_full_cycle("FAST")

    # Quick tshark capture for fast mode
    if g_tshark_path:
        print(f"\n  {CYAN}Capturing 15s network sample with tshark...{RESET}")
        fast_rows = []
        fast_done = threading.Event()
        t = threading.Thread(target=capture_tshark_background,
                             args=("FAST", 15, fast_rows, fast_done), daemon=True)
        t.start()
        fast_done.wait(timeout=30)
        if fast_rows:
            save_traffic_to_excel(fast_rows)
            print(f"  {GREEN}✔ {len(fast_rows)} packets captured → Network Traffic sheet.{RESET}")

    print(f"\n  {GREEN}{'═'*60}{RESET}")
    print(f"  {MAGENTA}⚡ YOUR REPORT IS READY!{RESET}")
    print(f"  {GREEN}  {g_excel_path}{RESET}")
    print(f"  {GREEN}{'═'*60}{RESET}\n")


# ─── MONITORING LOOP ──────────────────────────────────────────────────────────
def monitoring_loop():
    start_time   = time.time()
    duration_end = (start_time + g_duration) if g_duration > 0 else None
    cycle_num    = 1

    print(f"\n  {WHITE}{'─'*60}{RESET}")
    print(f"  {GREEN} IT Angel is now protecting this system. Ctrl+C to stop.{RESET}")
    print(f"  {WHITE}{'─'*60}{RESET}")

    try:
        while not g_stop_event.is_set():
            if duration_end and time.time() >= duration_end:
                print(f"\n  {YELLOW}⏱ Protection period complete.{RESET}")
                break

            run_full_cycle(cycle_num)
            cycle_num += 1

            if duration_end and time.time() >= duration_end:
                break

            show_live_traffic_and_countdown(CYCLE_INTERVAL, cycle_num-1, duration_end)

    except KeyboardInterrupt:
        pass

    print(f"\n  {GREEN}{'═'*60}{RESET}")
    print(f"  {WHITE} Session complete. {cycle_num-1} cycle(s) executed.{RESET}")
    print(f"  {GREEN} Report saved to:{RESET}")
    print(f"  {CYAN}  {g_excel_path}{RESET}")
    print(f"  {GREEN}{'═'*60}{RESET}\n")


# ─── MAIN MENU ────────────────────────────────────────────────────────────────
def main():
    while True:
        print_banner()

        # Detect tshark
        if detect_tshark():
            print(f"  {GREEN}✔ tshark detected:{RESET} {g_tshark_path}\n")
        else:
            print(f"  {YELLOW}⚠ tshark not found. Live traffic monitoring will be disabled.")
            print(f"    Install Wireshark (Windows) or: sudo apt install tshark (Linux){RESET}\n")

        select_os()
        select_duration()

        if g_duration == -1:
            # Fast check mode
            setup_excel()
            run_fast_check()
            print(f"  {WHITE}Press Enter to run another check or Ctrl+C to exit.{RESET}")
            try:
                input()
            except KeyboardInterrupt:
                break
            continue

        # Normal / Indefinite monitoring
        setup_excel()
        run_baseline()
        monitoring_loop()

        print(f"  {WHITE}Press Enter to start a new session or Ctrl+C to exit.{RESET}")
        try:
            input()
        except KeyboardInterrupt:
            break


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print(f"\n\n  {YELLOW}IT Angel stopped. Stay protected.{RESET}\n")
